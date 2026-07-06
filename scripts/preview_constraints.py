#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Compare a raw constraints xlsx with the CLEANED / processed pipeline output.

This is the offline "before vs after" view: it prints the raw spreadsheet grid
and, right below it, the cleaned result the import pipeline produces — guards
grouped, shift cells decoded, overlapping windows **merged** (union hours, not
the sum), notes preserved, and any non-blocking parse errors surfaced. Use it to
eyeball that the demo workbook comes out clean before importing for real.

Run (no server needed):
    python3 scripts/preview_constraints.py                       # demo file
    python3 scripts/preview_constraints.py אילוצים_דמו_מלא.xlsx   # any file
    python3 scripts/preview_constraints.py path/to/file.xlsx --raw   # also dump raw grid

It self-resolves ``backend/.venv`` so the bare system python works too.
"""
from __future__ import annotations

import os
import subprocess
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
BACKEND = REPO_ROOT / "backend"
DEFAULT_FILE = REPO_ROOT / "אילוצים_דמו_מלא.xlsx"


def _reexec_in_backend_venv() -> None:
    """Re-run this script under backend/.venv if the app deps aren't importable.

    Probe only ``fastapi`` (a heavy app dep) — importing app packages here would
    trigger their eager side-effect imports before the path is even set up.
    """
    try:
        import fastapi  # noqa: F401
        return
    except Exception:
        pass
    venv_py = BACKEND / ".venv" / "bin" / "python"
    # NB: don't resolve() — the venv python is a symlink to the system python, so
    # resolving both sides makes them compare equal and the re-exec never fires.
    if venv_py.exists() and Path(sys.executable) != venv_py:
        env = dict(os.environ)
        env["PYTHONPATH"] = str(BACKEND) + os.pathsep + env.get("PYTHONPATH", "")
        os.execve(str(venv_py), [str(venv_py), __file__, *sys.argv[1:]], env)


_reexec_in_backend_venv()
sys.path.insert(0, str(BACKEND))

# Importing app.services eagerly constructs app.config.Settings(), which needs a
# few env vars. Provide harmless defaults (mirrors tests/conftest.py) so this
# read-only tool runs without a configured .env.
for _key, _val in {
    "DATABASE_URL": "sqlite+aiosqlite://",
    "TELEGRAM_BOT_TOKEN": "preview-token",
    "APP_URL": "http://localhost:3000",
    "ADMIN_API_KEY": "preview-admin-key",
    "JWT_SECRET_KEY": "preview-jwt-secret",
    "ENVIRONMENT": "dev",
}.items():
    os.environ.setdefault(_key, _val)

from app.services.constraints_import.parser import parse_constraints_xlsx  # noqa: E402
from app.services.constraints_import.preview import build_preview  # noqa: E402

DAY_NAMES = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת"]


def _print_raw(path: Path) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["אילוצים"] if "אילוצים" in wb.sheetnames else wb.active
    print(f"\n=== RAW GRID · sheet '{ws.title}' ({ws.max_row}×{ws.max_column}) ===")
    for r in range(1, ws.max_row + 1):
        cells = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            cells.append("" if v is None else str(v))
        print(f"{r:>3} | " + " | ".join(cells))
    wb.close()


def _print_clean(path: Path) -> None:
    parsed = parse_constraints_xlsx(path.read_bytes())
    preview = build_preview(parsed, existing_names=set())

    print("\n=== CLEANED / PROCESSED OUTPUT ===")
    print(f"שבוע יעד: {preview.week_start} עד {preview.week_end}")
    print(f"מאבטחים: {len(preview.guards)}")

    if preview.errors:
        print(f"\n⚠️  שגיאות פרסור ({len(preview.errors)}):")
        for e in preview.errors:
            print(f"    - {e}")
    else:
        print("✓ אין שגיאות פרסור")

    for g in preview.guards:
        flag = "קיים" if g.exists else "חדש"
        print(f"\n● {g.name}  [{flag}]  · {g.weekly_hours} שעות/שבוע")
        if g.notes:
            print(f"    הערה: {g.notes}")
        for day in g.days:
            windows = ", ".join(day.segments) if day.segments else "—"
            hours = f"{day.hours}ש'" if day.hours else ""
            print(f"    {day.day_name:<6} {windows:<28} {hours}")


def main() -> int:
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    show_raw = "--raw" in sys.argv[1:]
    path = Path(args[0]) if args else DEFAULT_FILE
    if not path.is_absolute():
        path = (REPO_ROOT / path) if not path.exists() else path
    if not path.exists():
        print(f"קובץ לא נמצא: {path}", file=sys.stderr)
        return 1

    print(f"קובץ: {path}")
    if show_raw:
        _print_raw(path)
    _print_clean(path)
    print("\n(להצגת הגריד הגולמי לצד הפלט: הוסף --raw)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
