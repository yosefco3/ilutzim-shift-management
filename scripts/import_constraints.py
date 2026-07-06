#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""End-to-end constraints pipeline: an input xlsx → the availability model.

This is the single "pipe" that runs the whole import flow from one command:

    input .xlsx  →  parse  →  union-merge  →  WeeklySubmission/DailyStatus/ShiftWindow

The availability model is the **single source of truth** the schedule builder
reads from, so this takes the data as far as it can go today (the visual board
itself — Stage-B tasks 02–11 — isn't built yet).

──────────────────────────────────────────────────────────────────────────────
THE ONE KNOB TO FLIP FOR THE REAL FILE
──────────────────────────────────────────────────────────────────────────────
Right now the input defaults to the generated DEMO workbook. When Stage A's
real constraints export is ready, either:
  * pass its path as the first argument, or
  * set the env var  CONSTRAINTS_INPUT=/path/to/real.xlsx
…and then delete the ``DEMO_INPUT`` default + the ``# TODO(stage-a)`` marker
below. Nothing else in the pipeline is demo-specific — the parser/endpoints
already accept any correctly-formatted workbook.

──────────────────────────────────────────────────────────────────────────────
USAGE
──────────────────────────────────────────────────────────────────────────────
    python3 scripts/import_constraints.py                  # dry-run on the demo file (no DB)
    python3 scripts/import_constraints.py --raw            # dry-run + raw grid
    python3 scripts/import_constraints.py --commit         # WRITE into the availability model
    python3 scripts/import_constraints.py --commit --create-week   # also create the target week if missing
    python3 scripts/import_constraints.py /path/real.xlsx --commit  # the real Stage-A file

Dry-run is the default and never touches the database. ``--commit`` writes to
the DB configured in ``backend/.env`` (real PostgreSQL).
"""
from __future__ import annotations

import argparse
import asyncio
import os
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
BACKEND = REPO_ROOT / "backend"

# ── INPUT SOURCE — the one knob to flip for the real Stage-A file ─────────────
# TODO(stage-a): demo workbook for now. Point CONSTRAINTS_INPUT (or the CLI arg)
# at Stage A's real export when ready, then remove this default.
DEMO_INPUT = REPO_ROOT / "אילוצים_דמו_מלא.xlsx"


def _reexec_in_backend_venv() -> None:
    """Re-run under backend/.venv if the app deps aren't importable.

    The venv python is a symlink to the system python, so compare *unresolved*
    paths — resolving both sides makes them compare equal and the guard never
    fires.
    """
    try:
        import fastapi  # noqa: F401
        return
    except Exception:
        pass
    venv_py = BACKEND / ".venv" / "bin" / "python"
    if venv_py.exists() and Path(sys.executable) != venv_py:
        env = dict(os.environ)
        env["PYTHONPATH"] = str(BACKEND) + os.pathsep + env.get("PYTHONPATH", "")
        os.execve(str(venv_py), [str(venv_py), __file__, *sys.argv[1:]], env)


def _resolve_input(arg_path: str | None) -> Path:
    raw = arg_path or os.environ.get("CONSTRAINTS_INPUT") or str(DEMO_INPUT)
    path = Path(raw)
    if not path.is_absolute():
        candidate = REPO_ROOT / path
        path = candidate if candidate.exists() else path.resolve()
    return path


def main() -> int:
    parser = argparse.ArgumentParser(description="Constraints import pipeline")
    parser.add_argument("input", nargs="?", help="input xlsx (default: demo / CONSTRAINTS_INPUT)")
    parser.add_argument("--commit", action="store_true", help="write into the availability model")
    parser.add_argument("--create-week", action="store_true", help="create the target week if missing")
    parser.add_argument("--raw", action="store_true", help="also print the raw spreadsheet grid")
    parser.add_argument("--yes", action="store_true", help="skip the demo-file commit confirmation")
    args = parser.parse_args()

    input_path = _resolve_input(args.input)
    if not input_path.exists():
        print(f"קובץ קלט לא נמצא: {input_path}", file=sys.stderr)
        return 1

    _reexec_in_backend_venv()
    # chdir into backend so app.config picks up backend/.env (real DB creds).
    os.chdir(BACKEND)
    sys.path.insert(0, str(BACKEND))

    # Quiet SQLAlchemy's statement echo (ENVIRONMENT=dev turns it on) — this is a
    # human-facing CLI, not the app server.
    import logging
    logging.getLogger("sqlalchemy.engine").setLevel(logging.WARNING)

    from app.services.constraints_import.parser import parse_constraints_xlsx
    from app.services.constraints_import.preview import build_preview

    is_demo = input_path.resolve() == DEMO_INPUT.resolve()
    print(f"קלט: {input_path}" + ("   [קובץ דמו — להחלפה בקובץ שלב א']" if is_demo else ""))

    data = input_path.read_bytes()
    parsed = parse_constraints_xlsx(data)

    if args.raw:
        _print_raw(input_path)

    if not args.commit:
        _print_clean(build_preview(parsed, existing_names=set()))
        print("\n— הרצת יבש (dry-run). שום דבר לא נכתב. הוסף --commit כדי לשמור למודל הזמינות. —")
        return 0

    # Footgun guard: don't pollute a real DB with the DEMO file by accident.
    db_url = os.environ.get("DATABASE_URL", "")
    if is_demo and not db_url.startswith("sqlite") and not args.yes:
        print("\n⚠️  עומד לכתוב את קובץ ה־**דמו** ל-DB האמיתי (לא sqlite).")
        if sys.stdin.isatty():
            if input("    להמשיך? הקלד 'yes': ").strip().lower() not in ("yes", "y"):
                print("בוטל.")
                return 0
        else:
            print("    הרצה לא-אינטראקטיבית — הוסף --yes כדי לאשר. בוטל.")
            return 0

    return asyncio.run(_commit(parsed, create_week=args.create_week))


async def _commit(parsed, *, create_week: bool) -> int:
    from app.constants import WeekStatus
    from app.database import get_session
    from app.models.schedule_week import ScheduleWeek
    from app.repositories.schedule_week_repository import ScheduleWeekRepository
    from app.repositories.submission_repository import SubmissionRepository
    from app.repositories.user_repository import UserRepository
    from app.services.constraints_import.commit import (
        ConstraintsCommitService,
        WeekNotFoundError,
    )
    from app.services.submission_service import SubmissionService

    async with get_session() as session:
        week_repo = ScheduleWeekRepository(session)

        # Optionally seed the target week (the import never creates one silently).
        if create_week and parsed.week_start and parsed.week_end:
            existing = await week_repo.get_by_date_range(parsed.week_start, parsed.week_end)
            if existing is None:
                week = ScheduleWeek(
                    start_date=parsed.week_start,
                    end_date=parsed.week_end,
                    status=WeekStatus.CLOSED,
                )
                await week_repo.save(week)
                print(f"נוצר שבוע יעד חדש: {parsed.week_start}–{parsed.week_end} (closed)")

        commit_service = ConstraintsCommitService(
            user_repo=UserRepository(session),
            week_repo=ScheduleWeekRepository(session),
            submission_service=SubmissionService(
                submission_repo=SubmissionRepository(session),
                user_repo=UserRepository(session),
                week_repo=ScheduleWeekRepository(session),
            ),
        )
        try:
            resp = await commit_service.commit(parsed)
        except WeekNotFoundError as exc:
            print(f"\n❌ {exc}", file=sys.stderr)
            print("   טיפ: הוסף --create-week כדי ליצור את השבוע אוטומטית.", file=sys.stderr)
            return 2

    s = resp.summary
    print("\n=== סיכום ייבוא (נשמר למודל הזמינות) ===")
    print(f"שבוע יעד: {s.week_start} עד {s.week_end}")
    print(f"✅ {s.imported} מאבטחים יובאו ונשמרו")
    print(f"➕ מתוכם {s.created_new} נוצרו כמאבטחים חדשים")
    if s.errors:
        print(f"⚠️  {len(s.errors)} שגיאות:")
        for e in s.errors:
            print(f"    - {e}")
    print("\nהמאבטחים והזמינות שלהם נראים עכשיו במסך \"הגשות\" לשבוע היעד.")
    return 0


# ── pretty printers (shared shape with scripts/preview_constraints.py) ────────

def _print_raw(path: Path) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["אילוצים"] if "אילוצים" in wb.sheetnames else wb.active
    print(f"\n=== RAW GRID · sheet '{ws.title}' ({ws.max_row}×{ws.max_column}) ===")
    for r in range(1, ws.max_row + 1):
        cells = ["" if (v := ws.cell(r, c).value) is None else str(v)
                 for c in range(1, ws.max_column + 1)]
        print(f"{r:>3} | " + " | ".join(cells))
    wb.close()


def _print_clean(preview) -> None:
    print("\n=== CLEANED / PROCESSED OUTPUT ===")
    print(f"שבוע יעד: {preview.week_start} עד {preview.week_end}")
    print(f"מאבטחים: {len(preview.guards)}")
    if preview.errors:
        print(f"\n⚠️  שגיאות פרסור ({len(preview.errors)}):")
        for e in preview.errors:
            print(f"    - {e}")
    else:
        print("✓ אין שגיאות פרסור")
    for g in preview.guards:
        flag = "קיים" if g.exists else "חדש"
        print(f"\n● {g.name}  [{flag}]  · {g.weekly_hours} שעות/שבוע")
        if g.notes:
            print(f"    הערה: {g.notes}")
        for day in g.days:
            windows = ", ".join(day.segments) if day.segments else "—"
            hours = f"{day.hours}ש'" if day.hours else ""
            print(f"    {day.day_name:<6} {windows:<28} {hours}")


if __name__ == "__main__":
    raise SystemExit(main())
