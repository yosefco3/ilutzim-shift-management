"""Regenerate דוגמה_אילוצים_מאבטחים.xlsx — the sample constraints workbook.

The static sample shipped at the repo root mirrors the output of
``ExcelExportService.export_constraints_report``. This script rebuilds it from
the hard-coded demo data below, reusing the same style constants and the
``_draw_guard_separator`` helper so the sample always reflects the real export
look (including the thick line drawn between guards).

Run with the backend venv so openpyxl resolves:
    backend/.venv/bin/python scripts/generate_constraints_example.py
"""

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT / "backend"))

from app.services.excel_export_service import (  # noqa: E402
    _CENTER,
    _CENTER_WRAP,
    _DAY_NAMES_HE,
    _EMPTY_FILL,
    _GREEN_FILL,
    _RED_FILL,
    _SHIFT_PERIODS,
    _THIN_BORDER,
    _TITLE_FONT,
    _apply_header_style,
    _draw_guard_separator,
    _merge_vertical,
)

OUTPUT = _ROOT / "דוגמה_אילוצים_מאבטחים.xlsx"

WEEK_RANGE = ("2026-06-14", "2026-06-20")

# Period keys in display order.
_M, _E, _N = (p[0] for p in _SHIFT_PERIODS)  # morning / afternoon / night values
# Friendlier local aliases keyed by the Hebrew labels used in the demo data.
B, V, L = "בוקר", "ערב", "לילה"
_LABEL_TO_TYPE = {B: _M, V: _E, L: _N}

# Per-guard demo data. Each "day" is one of:
#   "unavailable"           → merged red "לא זמין"
#   "available"             → merged green "זמין" (available, no windows)
#   {period: "HH:MM–HH:MM"} → per-period time windows (missing periods = orange)
GUARDS = [
    {
        "name": "אבי כהן",
        "phone": "050-1234567",
        "notes": "מעדיף משמרות בוקר",
        "days": [
            {B: "07:00–16:00"},
            {B: "07:00–16:00"},
            {B: "07:00–13:00", V: "15:00–19:00"},
            {B: "07:00–16:00"},
            {B: "07:00–16:00"},
            "unavailable",
            "unavailable",
        ],
    },
    {
        "name": "בני לוי",
        "phone": "052-2345678",
        "notes": "",
        "days": [
            {V: "15:00–23:00", L: "23:00–07:00"},
            {L: "23:00–07:00"},
            "unavailable",
            {V: "15:00–23:00"},
            {V: "15:00–23:00", L: "23:00–07:00"},
            {L: "23:00–07:00"},
            {L: "23:00–07:00"},
        ],
    },
    {
        "name": "גדי מזרחי",
        "phone": "054-3456789",
        "notes": "מילואים בתחילת השבוע",
        "days": [
            "unavailable",
            "unavailable",
            {B: "08:00–16:00"},
            {B: "08:00–16:00"},
            {B: "08:00–16:00", V: "16:00–22:00"},
            {B: "08:00–13:00"},
            "unavailable",
        ],
    },
    {
        "name": "דנה אזולאי",
        "phone": "053-4567890",
        "notes": "זמינה בכל המשמרות",
        "days": [
            {B: "07:00–15:00", V: "15:00–23:00"},
            {B: "07:00–15:00"},
            {V: "15:00–23:00"},
            {B: "07:00–15:00", L: "23:00–07:00"},
            {V: "15:00–23:00"},
            {B: "07:00–14:00"},
            "available",
        ],
    },
    {
        "name": "הראל ביטון",
        "phone": "058-5678901",
        "notes": "",
        "days": [
            {V: "16:00–23:00"},
            "unavailable",
            "unavailable",
            {B: "09:00–17:00"},
            "unavailable",
            "unavailable",
            {V: "17:00–23:00"},
        ],
    },
]

_PERIOD_COL = 3
_FIRST_DAY_COL = 4
_NOTES_COL = 11
N_COLS = 11
_SPAN = len(_SHIFT_PERIODS)


def build() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "אילוצים"
    ws.sheet_view.rightToLeft = True

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COLS)
    title = ws.cell(
        row=1, column=1, value=f"אילוצים שהוגשו — {WEEK_RANGE[0]} עד {WEEK_RANGE[1]}"
    )
    title.font = _TITLE_FONT
    title.alignment = _CENTER

    # Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N_COLS)
    ws.cell(
        row=2, column=1, value=f"סך הכל הגישו: {len(GUARDS)} מאבטחים"
    ).alignment = _CENTER

    # Headers
    headers = ["שם", "טלפון", "משמרת"] + _DAY_NAMES_HE + ["הערות"]
    for col, header in enumerate(headers, 1):
        _apply_header_style(ws.cell(row=4, column=col, value=header))

    row_num = 5
    for guard in GUARDS:
        _merge_vertical(ws, row_num, 1, _SPAN, guard["name"], thick_bottom=True)
        _merge_vertical(ws, row_num, 2, _SPAN, guard["phone"], thick_bottom=True)
        _merge_vertical(
            ws,
            row_num,
            _NOTES_COL,
            _SPAN,
            guard["notes"],
            alignment=Alignment(
                horizontal="right", vertical="center", wrap_text=True
            ),
            thick_bottom=True,
        )

        # Period label column
        for p, (_type, label, accent) in enumerate(_SHIFT_PERIODS):
            cell = ws.cell(row=row_num + p, column=_PERIOD_COL, value=label)
            cell.border = _THIN_BORDER
            cell.alignment = _CENTER
            cell.fill = accent
            cell.font = Font(bold=True)

        for day_offset, day in enumerate(guard["days"]):
            col = _FIRST_DAY_COL + day_offset

            if day == "unavailable":
                _merge_vertical(
                    ws, row_num, col, _SPAN, "לא זמין",
                    fill=_RED_FILL, thick_bottom=True,
                )
                continue
            if day == "available":
                _merge_vertical(
                    ws, row_num, col, _SPAN, "זמין",
                    fill=_GREEN_FILL, thick_bottom=True,
                )
                continue

            for p, (shift_type, _label, _accent) in enumerate(_SHIFT_PERIODS):
                cell = ws.cell(row=row_num + p, column=col)
                cell.border = _THIN_BORDER
                cell.alignment = _CENTER_WRAP
                value = None
                for he_label, win in day.items():
                    if _LABEL_TO_TYPE[he_label] == shift_type:
                        value = win
                        break
                if value:
                    cell.value = value
                    cell.fill = _GREEN_FILL
                else:
                    cell.fill = _EMPTY_FILL

        _draw_guard_separator(ws, row_num + _SPAN - 1, N_COLS)
        row_num += _SPAN

    # Column widths (match the real export)
    ws.column_dimensions[get_column_letter(1)].width = 20
    ws.column_dimensions[get_column_letter(2)].width = 16
    ws.column_dimensions[get_column_letter(_PERIOD_COL)].width = 10
    for col in range(_FIRST_DAY_COL, _NOTES_COL):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions[get_column_letter(_NOTES_COL)].width = 30

    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    build()
