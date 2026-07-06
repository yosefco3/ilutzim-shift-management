#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate a realistic DEMO constraints workbook (אילוצים) for Stage-B testing.

Mirrors the format/styling of ``דוגמה_אילוצים_מאבטחים.xlsx`` (which itself mirrors
the real export in ``app/services/excel_export_service.py``):

  * sheet "אילוצים"; title + "סך הכל הגישו" rows; blue header row
  * 3 rows per guard (בוקר/ערב/לילה); name/phone/notes merged across the 3 rows
  * per day: a full availability window (green), an off-shift on an otherwise
    available day (orange), a whole-day "לא זמין" (red, merged), or "זמין" (green,
    merged = available all day)
  * en-dash (–, U+2013) in time ranges

Realistic ratios (configurable below): most guards prefer morning, then evening,
few night; most are NOT armed. A handful of guards have overlapping morning+evening
windows so the union-hours rule (07:00–16:30 ∪ 15:00–19:00 = 12h, not 13.5) is
exercised by the pipeline.

Run:  backend/.venv/bin/python scripts/generate_demo_constraints.py
"""
from __future__ import annotations

import random
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── config ──────────────────────────────────────────────────────────────
N_GUARDS = 36
OUT_PATH = "אילוצים_דמו_מלא.xlsx"
SEED = 42


def _next_week_title(today: date | None = None) -> str:
    """The upcoming planning week (Sunday→Saturday, strictly after this week)."""
    today = today or date.today()
    current_start = today - timedelta(days=(today.weekday() + 1) % 7)  # Sunday
    start = current_start + timedelta(days=7)
    end = start + timedelta(days=6)
    return f"אילוצים שהוגשו — {start.isoformat()} עד {end.isoformat()}"


WEEK_TITLE = _next_week_title()

DASH = "–"  # en-dash, matches the example/export
DAYS = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת"]  # 0..6
SHIFTS = ["בוקר", "ערב", "לילה"]  # rows per guard

# ── styling (hex copied from excel_export_service._SHIFT_PERIODS etc.) ───
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
GREEN = PatternFill("solid", fgColor="C6EFCE")   # an availability window / "זמין"
RED = PatternFill("solid", fgColor="FF4444")     # whole-day "לא זמין"
ORANGE = PatternFill("solid", fgColor="FFC000")  # available day, this shift not chosen
SHIFT_FILL = {
    "בוקר": PatternFill("solid", fgColor="FFF2CC"),
    "ערב": PatternFill("solid", fgColor="FCE4D6"),
    "לילה": PatternFill("solid", fgColor="D9E1F2"),
}
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
THICK = Side(style="medium", color="808080")

# ── name pools (paired index-wise → unique full names) ──────────────────
FIRST = [
    "אבי", "בני", "גדי", "דני", "הראל", "יוסי", "משה", "דוד", "יעקב", "איתי",
    "רון", "עומר", "ניר", "גיא", "אורי", "טל", "עידן", "שגיא", "ליאור", "נדב",
    "אסף", "רועי", "יובל", "אלון", "עידו", "שי", "דור", "עמית", "נועם", "אייל",
    "ישי", "מתן", "אדם", "רן", "יואב", "אלעד",
]
LAST = [
    "כהן", "לוי", "מזרחי", "אזולאי", "ביטון", "פרץ", "דהן", "אוחיון", "גבאי", "חדד",
    "ועקנין", "אברהם", "מלכה", "יוסף", "בן דוד", "שמעוני", "אדרי", "סבן", "נחום", "רוזן",
    "שרון", "אטיאס", "אסולין", "עמר", "חזן", "בן חמו", "צרפתי", " נגר", "ברק", "שלום",
    "טובול", "אוחנה", "בוזגלו", "סויסה", "אלבז", "ממן",
]

MORNING = [("07:00", "15:00"), ("07:00", "16:00"), ("08:00", "16:00"),
           ("07:30", "15:30"), ("07:00", "14:30"), ("09:00", "17:00")]
MORNING_SHORT = [("07:00", "13:00"), ("07:00", "14:00"), ("08:00", "13:00")]
EVENING = [("15:00", "23:00"), ("16:00", "23:00"), ("17:00", "23:00"), ("15:00", "22:00")]
NIGHT = [("23:00", "07:00"), ("19:00", "07:00"), ("16:00", "07:00"), ("22:00", "06:00")]
OVL_M, OVL_E = ("07:00", "16:30"), ("15:00", "19:00")  # union 07:00–19:00 = 12h

NOTES_MORNING = ["מעדיף משמרות בוקר", "זמין בעיקר בוקר", "לא זמין בסופ\"ש",
                 "מילואים בתחילת השבוע", "יש רכב עירייה"]
NOTES_EVENING = ["מעדיף ערב", "לא זמין בבקרים", "זמין מהצהריים"]
NOTES_NIGHT = ["עושה לילות", "זמין ללילות בלבד", "מעדיף לילה"]
NOTES_ARMED = ["אחמ\"ש / חמוש", "חמוש"]


# Every submitted window must sit inside the operational day 07:00 → 07:00
# next morning: no starts before 07:00, no ends past 07:00 (e.g. 06:30–…,
# 23:00–07:30 are both invalid submissions).
_DAY_START_MIN = 7 * 60


def _minutes(hhmm):
    h, m = (int(x) for x in hhmm.split(":"))
    return h * 60 + m


def win(pair):
    start, end = _minutes(pair[0]), _minutes(pair[1])
    if end <= start:  # crosses midnight
        end += 24 * 60
    assert _DAY_START_MIN <= start and end <= _DAY_START_MIN + 24 * 60, (
        f"חלון {pair[0]}–{pair[1]} חורג מגבול 07:00→07:00"
    )
    return f"{pair[0]}{DASH}{pair[1]}"


# cell kinds returned by plan_day(): one of these
UNAVAIL = "UNAVAIL"      # whole day off → merged "לא זמין" red
ALLDAY = "ALLDAY"        # available all day → merged "זמין" green
# else: dict {shift_label: window_str} for the shifts that have a window


def plan_guard(rng, category):
    """Return (days, note) where days[d] is UNAVAIL / ALLDAY / {shift: window}."""
    days = {}
    for d in range(7):
        days[d] = plan_day(rng, category, d)
    note = pick_note(rng, category)
    return days, note


def plan_day(rng, category, d):
    weekday = d <= 4          # Sun–Thu
    friday = d == 5
    saturday = d == 6

    if category == "night":
        if weekday:
            if rng.random() < 0.78:
                return {"לילה": win(rng.choice(NIGHT))}
            return UNAVAIL
        if friday:
            return {"לילה": win(rng.choice(NIGHT))} if rng.random() < 0.4 else UNAVAIL
        return {"לילה": win(rng.choice(NIGHT))} if rng.random() < 0.25 else UNAVAIL

    if category == "evening":
        if weekday:
            r = rng.random()
            if r < 0.05:
                return UNAVAIL
            shifts = {"ערב": win(rng.choice(EVENING))}
            if rng.random() < 0.15:          # also some morning
                shifts["בוקר"] = win(rng.choice(MORNING))
            return shifts
        if friday:
            if rng.random() < 0.5:
                return UNAVAIL
            return {"ערב": win(("16:00", "21:00"))}
        return UNAVAIL if rng.random() < 0.85 else {"ערב": win(("16:00", "22:00"))}

    # category == "morning" (the majority)
    if weekday:
        r = rng.random()
        if r < 0.08:
            return UNAVAIL
        if r < 0.11:
            return ALLDAY                     # rare full availability
        # overlapping morning+evening for a few → exercises union-hours rule
        if rng.random() < 0.18:
            return {"בוקר": win(OVL_M), "ערב": win(OVL_E)}
        shifts = {"בוקר": win(rng.choice(MORNING))}
        if rng.random() < 0.12:               # occasional extra evening (non-overlap)
            shifts["ערב"] = win(("17:00", "21:00"))
        return shifts
    if friday:
        r = rng.random()
        if r < 0.5:
            return {"בוקר": win(rng.choice(MORNING_SHORT))}
        return UNAVAIL
    # saturday
    r = rng.random()
    if r < 0.8:
        return UNAVAIL
    if r < 0.9:
        return ALLDAY
    return {"בוקר": win(("07:30", "14:00"))}


def pick_note(rng, category):
    pool = {"morning": NOTES_MORNING, "evening": NOTES_EVENING, "night": NOTES_NIGHT}[category]
    parts = []
    if rng.random() < 0.45:
        parts.append(rng.choice(pool))
    if rng.random() < 0.14:                    # few armed
        parts.append(rng.choice(NOTES_ARMED))
    return " · ".join(parts) if parts else None


def build():
    rng = random.Random(SEED)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "אילוצים"
    ws.sheet_view.rightToLeft = True

    # categories with the requested ratios (morning ≫ evening > night)
    n_morning = round(N_GUARDS * 0.61)
    n_evening = round(N_GUARDS * 0.25)
    n_night = N_GUARDS - n_morning - n_evening
    categories = (["morning"] * n_morning + ["evening"] * n_evening + ["night"] * n_night)
    rng.shuffle(categories)

    # header block
    ws["A1"] = WEEK_TITLE
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = f"סך הכל הגישו: {N_GUARDS} מאבטחים"
    ws.merge_cells("A1:K1")
    ws.merge_cells("A2:K2")
    headers = ["שם", "טלפון", "משמרת", *DAYS, "הערות"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(4, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    widths = {"A": 20, "B": 16, "C": 10, "K": 30}
    for col in "DEFGHIJ":
        widths[col] = 14
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    row = 5
    for i in range(N_GUARDS):
        cat = categories[i]
        name = f"{FIRST[i]} {LAST[i]}".strip()
        phone = f"05{rng.choice([0, 2, 3, 4, 8])}-{rng.randint(1000000, 9999999)}"
        days, note = plan_guard(rng, cat)
        top, bot = row, row + 2

        # name / phone / notes merged across the 3 shift rows
        ws.cell(top, 1, name)
        ws.cell(top, 2, phone)
        if note:
            ws.cell(top, 11, note)
        for col in (1, 2, 11):
            ws.merge_cells(start_row=top, start_column=col, end_row=bot, end_column=col)
            ws.cell(top, col).alignment = CENTER

        # shift labels column C
        for p, label in enumerate(SHIFTS):
            cell = ws.cell(top + p, 3, label)
            cell.font = Font(bold=True)
            cell.fill = SHIFT_FILL[label]
            cell.alignment = CENTER

        # day columns D..J (4..10)
        for d in range(7):
            col = 4 + d
            plan = days[d]
            if plan in (UNAVAIL, ALLDAY):
                text = "לא זמין" if plan == UNAVAIL else "זמין"
                fill = RED if plan == UNAVAIL else GREEN
                ws.cell(top, col, text)
                ws.merge_cells(start_row=top, start_column=col, end_row=bot, end_column=col)
                tl = ws.cell(top, col)
                tl.alignment = CENTER
                tl.fill = fill
            else:  # dict of windows
                for p, label in enumerate(SHIFTS):
                    cell = ws.cell(top + p, col)
                    cell.alignment = CENTER
                    if label in plan:
                        cell.value = plan[label]
                        cell.fill = GREEN
                    else:
                        cell.fill = ORANGE  # available day, this shift not chosen

        # borders: thin inside, medium under each guard block
        for r in range(top, bot + 1):
            for c in range(1, 12):
                bottom = THICK if r == bot else THIN
                ws.cell(r, c).border = Border(left=THIN, right=THIN, top=THIN, bottom=bottom)

        row = bot + 1

    wb.save(OUT_PATH)
    print(f"wrote {OUT_PATH}: {N_GUARDS} guards "
          f"(morning={n_morning}, evening={n_evening}, night={n_night})")


if __name__ == "__main__":
    build()
