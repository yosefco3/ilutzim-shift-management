"""Pure parser for the guard-constraints xlsx.

Input  = raw ``.xlsx`` bytes (the format emitted by the manager export, e.g.
``דוגמה_אילוצים_מאבטחים.xlsx``).
Output = a ``ParsedImport`` dataclass tree. **No DB access** — this layer is
fully deterministic and unit-tested against the real sample file.

Locked semantics (see ``STAGE_B_PROMPTS/README.md`` → "הכללים הנעולים"):

* Sheet ``אילוצים``. Row 1 = title with the week date-range, row 2 = total,
  row 3 = blank, row 4 = header ``שם|טלפון|משמרת|ראשון..שבת|הערות``.
  Then **3 rows per guard** (בוקר/ערב/לילה); name/phone/notes only on the
  first (morning) row.
* Shift mapping: בוקר→MORNING, **ערב→AFTERNOON**, לילה→NIGHT.
* Cell values: ``HH:MM–HH:MM`` window (accept en-dash *and* hyphen);
  ``זמין`` = all-day; ``לא זמין`` / empty = unavailable. ``end <= start`` on a
  window ⇒ wraps past midnight.
* Anything unparsable becomes an ``errors`` entry (with name+day+shift) and the
  cell falls back to UNAVAILABLE — the parser never raises on bad data.
"""

from __future__ import annotations

import io
import re
import uuid
from dataclasses import dataclass, field
from datetime import date, time
from enum import Enum

import openpyxl

from app.constants import ShiftType

from .attributes import split_notes


# --- returned data structure ------------------------------------------------

class CellKind(Enum):
    """What a single day×shift cell means."""
    WINDOW = "window"          # an explicit HH:MM–HH:MM range
    ALL_DAY = "all_day"        # the literal "זמין" — available the whole shift
    UNAVAILABLE = "unavailable"  # empty or "לא זמין"


@dataclass
class Cell:
    kind: CellKind
    start: time | None = None
    end: time | None = None
    wraps_midnight: bool = False


@dataclass
class ParsedGuard:
    name: str
    phone: str | None
    notes: str | None
    # Structured attributes (UserRole values) lifted out of the notes column —
    # e.g. ["AHMASH", "PATROL_VEHICLE"]. These are constraining factors the
    # board enforces; see ``attributes.split_notes``.
    roles: list[str] = field(default_factory=list)
    # The guard's DB id (UUID string) if the export carried a "מזהה" column —
    # lets the importer match the exact guard even if the name was edited.
    guard_id: str | None = None
    # cells[day_index 0..6][shift] -> Cell  (day 0 = ראשון … day 6 = שבת)
    cells: dict[int, dict[ShiftType, Cell]] = field(default_factory=dict)


@dataclass
class ParsedImport:
    week_start: date | None
    week_end: date | None
    guards: list[ParsedGuard] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


# --- parsing constants ------------------------------------------------------

_SHEET_NAME = "אילוצים"

# Header labels we anchor on (the header row is *found*, not assumed at row 4).
_HEADER_ID = "מזהה"
_HEADER_NAME = "שם"
_HEADER_PHONE = "טלפון"
_HEADER_SHIFT = "משמרת"

# Hebrew weekday header → day_index 0..6.
_DAY_HEADERS = {
    "ראשון": 0,
    "שני": 1,
    "שלישי": 2,
    "רביעי": 3,
    "חמישי": 4,
    "שישי": 5,
    "שבת": 6,
}

_NOTES_HEADER = "הערות"

# Shift label (normalised) → enum. We strip whitespace/quotes before lookup.
_SHIFT_LABELS = {
    "בוקר": ShiftType.MORNING,
    "ערב": ShiftType.AFTERNOON,
    "לילה": ShiftType.NIGHT,
}

_AVAILABLE_TOKEN = "זמין"
_UNAVAILABLE_TOKEN = "לא זמין"

# Anti-DoS caps on sheet dimensions (real files: 12 cols, ~3 rows per guard).
# Checked before iterating so an inflated max_row/max_column can't burn CPU.
_MAX_ROWS = 1000
_MAX_COLS = 60

# Accept en-dash (U+2013), em-dash, figure-dash and the plain hyphen.
_DASHES = "–—‒-"
_TIME_RANGE_RE = re.compile(
    r"^\s*(\d{1,2}:\d{2})\s*[" + _DASHES + r"]\s*(\d{1,2}:\d{2})\s*$"
)
# "… 2026-06-14 עד 2026-06-20 …"  → two ISO dates.
_DATE_RE = re.compile(r"(\d{4})-(\d{2})-(\d{2})")


def _norm(value) -> str:
    """Normalise a cell value to a trimmed string ('' for None)."""
    if value is None:
        return ""
    return str(value).replace("’", "").replace("'", "").strip()


def _parse_uuid(raw) -> str | None:
    """Return the canonical UUID string if ``raw`` is a valid UUID, else None."""
    text = _norm(raw)
    if not text:
        return None
    try:
        return str(uuid.UUID(text))
    except (ValueError, AttributeError, TypeError):
        return None


def _parse_time(token: str) -> time | None:
    try:
        hh, mm = token.split(":")
        h, m = int(hh), int(mm)
        if 0 <= h <= 23 and 0 <= m <= 59:
            return time(h, m)
    except (ValueError, AttributeError):
        pass
    return None


def _parse_cell(raw, ctx: str, errors: list[str]) -> Cell:
    """Decode one day×shift cell value into a ``Cell``.

    ``ctx`` is a human label ("<name> · <day> · <shift>") used in error text.
    """
    text = _norm(raw)
    if text == "" or text == _UNAVAILABLE_TOKEN:
        return Cell(CellKind.UNAVAILABLE)
    if text == _AVAILABLE_TOKEN:
        return Cell(CellKind.ALL_DAY)

    m = _TIME_RANGE_RE.match(text)
    if m:
        start = _parse_time(m.group(1))
        end = _parse_time(m.group(2))
        if start is not None and end is not None:
            return Cell(
                CellKind.WINDOW,
                start=start,
                end=end,
                wraps_midnight=end <= start,
            )

    errors.append(f"{ctx}: ערך לא תקין '{text}' → טופל כ'לא זמין'")
    return Cell(CellKind.UNAVAILABLE)


def _parse_week_range(title: str) -> tuple[date | None, date | None]:
    dates = _DATE_RE.findall(title or "")
    if len(dates) >= 2:
        try:
            start = date(int(dates[0][0]), int(dates[0][1]), int(dates[0][2]))
            end = date(int(dates[1][0]), int(dates[1][1]), int(dates[1][2]))
            return start, end
        except ValueError:
            return None, None
    return None, None


def _find_header_row(ws) -> int | None:
    """Locate the row that contains שם/טלפון/משמרת (don't assume row 4)."""
    for r in range(1, min(ws.max_row, 30) + 1):
        labels = {_norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)}
        if {_HEADER_NAME, _HEADER_PHONE, _HEADER_SHIFT} <= labels:
            return r
    return None


def parse_constraints_xlsx(data: bytes) -> ParsedImport:
    """Parse constraints xlsx bytes into a ``ParsedImport`` (no DB)."""
    errors: list[str] = []
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

    ws = wb[_SHEET_NAME] if _SHEET_NAME in wb.sheetnames else wb.active

    if ws.max_row > _MAX_ROWS or ws.max_column > _MAX_COLS:
        raise ValueError(
            f"workbook too large: {ws.max_row}x{ws.max_column} "
            f"(max {_MAX_ROWS}x{_MAX_COLS})"
        )

    # Week range from the title (first non-empty cell of row 1).
    title = ""
    for c in range(1, ws.max_column + 1):
        v = _norm(ws.cell(1, c).value)
        if v:
            title = v
            break
    week_start, week_end = _parse_week_range(title)

    header_row = _find_header_row(ws)
    if header_row is None:
        errors.append("לא נמצאה שורת כותרת (שם/טלפון/משמרת) — הקובץ אינו בפורמט הצפוי")
        return ParsedImport(week_start, week_end, [], errors)

    # Map columns from the header row.
    col_id = col_name = col_phone = col_shift = col_notes = None
    day_cols: dict[int, int] = {}  # day_index -> column
    for c in range(1, ws.max_column + 1):
        label = _norm(ws.cell(header_row, c).value)
        if label == _HEADER_ID:
            col_id = c
        elif label == _HEADER_NAME:
            col_name = c
        elif label == _HEADER_PHONE:
            col_phone = c
        elif label == _HEADER_SHIFT:
            col_shift = c
        elif label == _NOTES_HEADER:
            col_notes = c
        elif label in _DAY_HEADERS:
            day_cols[_DAY_HEADERS[label]] = c

    if col_name is None or col_shift is None or not day_cols:
        errors.append("שורת הכותרת חסרה עמודות חובה (שם/משמרת/ימים)")
        return ParsedImport(week_start, week_end, [], errors)

    guards: list[ParsedGuard] = []
    current: ParsedGuard | None = None

    for r in range(header_row + 1, ws.max_row + 1):
        name = _norm(ws.cell(r, col_name).value)
        shift_label = _norm(ws.cell(r, col_shift).value)

        # A new name in column A starts a new guard; id/phone/notes from here.
        if name:
            phone = _norm(ws.cell(r, col_phone).value) if col_phone else ""
            notes = _norm(ws.cell(r, col_notes).value) if col_notes else ""
            guard_id = _parse_uuid(ws.cell(r, col_id).value) if col_id else None
            # Lift attributes (אחמ"ש / רכב סיור / חמוש) out of the free-text
            # notes into structured roles; keep only the residual note text.
            roles, notes = split_notes(notes or None)
            current = ParsedGuard(
                name=name,
                phone=phone or None,
                notes=notes,
                roles=roles,
                guard_id=guard_id,
                cells={},
            )
            guards.append(current)

        if current is None:
            continue  # data before the first named guard — skip defensively

        shift = _SHIFT_LABELS.get(shift_label)
        if shift is None:
            if shift_label:
                errors.append(
                    f"{current.name}: משמרת לא מזוהה '{shift_label}' (שורה {r})"
                )
            continue

        for day_index, col in day_cols.items():
            raw = ws.cell(r, col).value
            day_he = next(k for k, v in _DAY_HEADERS.items() if v == day_index)
            ctx = f"{current.name} · {day_he} · {shift_label}"
            cell = _parse_cell(raw, ctx, errors)
            current.cells.setdefault(day_index, {})[shift] = cell

    wb.close()
    return ParsedImport(week_start, week_end, guards, errors)
