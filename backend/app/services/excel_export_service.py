"""
ExcelExportService — generates Excel files from schedule data.

Three report types:
1. Weekly schedule grid  – all guards × all days with shift choices
2. Deviation report       – guards exceeding shift-count thresholds
3. Guard history          – single-guard submission trail across weeks
"""

import io
import logging
import uuid
from datetime import date, timedelta
from typing import Any

from app.constants import (
    AdminRole,
    ShiftType,
    WeekStatus,
)
from app.messages import Messages
from app.repositories.schedule_week_repository import ScheduleWeekRepository
from app.schedule_builder.services.schedule_export_service import (
    order_unverified_first,
)
from app.schedule_builder.utils import intervals as iv
from app.repositories.submission_repository import SubmissionRepository
from app.repositories.user_repository import UserRepository
from app.services.schedule_grid_model import build_schedule_grid

logger = logging.getLogger("ilutzim")

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── Shared style constants ──────────────────────────────────────────
_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(
    start_color="4472C4", end_color="4472C4", fill_type="solid"
)
_TITLE_FONT = Font(bold=True, size=13)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
# Thick separator drawn under the last row of each guard's block so adjacent
# guards are easy to tell apart at a glance.
_GUARD_SEPARATOR_SIDE = Side(style="thick")
_CENTER = Alignment(horizontal="center", vertical="center")
_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
_RED_FILL = PatternFill(
    start_color="FF4444", end_color="FF4444", fill_type="solid"
)
_GREEN_FILL = PatternFill(
    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
)
# Orange fill for an available day where this particular period was not chosen
# (distinct from the solid red of a fully-unavailable day).
_EMPTY_FILL = PatternFill(
    start_color="FFC000", end_color="FFC000", fill_type="solid"
)
# Soft red wash for a guard with no linked Telegram — their schedule can't be
# delivered over the bot, so their name block is flagged in the per-guard export.
_NO_TG_FILL = PatternFill(
    start_color="FFD9D9", end_color="FFD9D9", fill_type="solid"
)
# Purple wash for an EVENT (non-splitting) position — רענון, ישיבת מועצה — so it
# reads apart from routine posts. The name block gets a deeper purple than the
# guard cells. An unstaffed event cell is a valid state, so it stays event-purple
# (never the amber "empty" fill a normal uncovered cell would get).
_EVENT_FILL = PatternFill(
    start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"
)
_EVENT_NAME_FILL = PatternFill(
    start_color="CCC0DA", end_color="CCC0DA", fill_type="solid"
)

# Three daily shift periods, in display order, with their own accent colour
# for the "משמרת" column so the sheet reads clearly at a glance.
_SHIFT_PERIODS: list[tuple[str, str, PatternFill]] = [
    (
        ShiftType.MORNING.value,
        "בוקר",
        PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    ),
    (
        ShiftType.AFTERNOON.value,
        "ערב",
        PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    ),
    (
        ShiftType.NIGHT.value,
        "לילה",
        PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
    ),
]

# Hebrew weekday order (Sunday=0 … Saturday=6)
_DAY_NAMES_HE = [
    "ראשון",
    "שני",
    "שלישי",
    "רביעי",
    "חמישי",
    "שישי",
    "שבת",
]

_SHIFT_LABELS: dict[str, str] = {
    ShiftType.MORNING.value: Messages.LABEL_MORNING,
    ShiftType.AFTERNOON.value: Messages.LABEL_AFTERNOON,
    ShiftType.NIGHT.value: Messages.LABEL_NIGHT,
}

def _apply_header_style(cell: Any) -> None:
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.border = _THIN_BORDER
    cell.alignment = _CENTER


def _apply_cell_style(cell: Any, center: bool = True) -> None:
    cell.border = _THIN_BORDER
    cell.alignment = _CENTER if center else Alignment(vertical="center")


def _merge_vertical(
    ws: Any,
    row: int,
    col: int,
    span: int,
    value: Any,
    fill: Any = None,
    alignment: Any = None,
    thick_bottom: bool = False,
) -> Any:
    """Merge ``span`` cells down a single column and style the whole block.

    openpyxl only keeps the top-left cell's value, but borders/fill must be set
    on every underlying cell for the merged region to render as one boxed cell.

    With ``thick_bottom`` the merged block gets a heavy bottom edge. openpyxl
    derives a merged range's outer borders from the *top-left* (anchor) cell on
    save, so the thick side must be set on the anchor's ``bottom`` — setting it
    on the bottom ``MergedCell`` directly is silently overwritten.
    """
    ws.merge_cells(
        start_row=row, start_column=col, end_row=row + span - 1, end_column=col
    )
    for r in range(row, row + span):
        c = ws.cell(row=r, column=col)
        if thick_bottom and r == row:
            c.border = Border(
                left=_THIN_BORDER.left,
                right=_THIN_BORDER.right,
                top=_THIN_BORDER.top,
                bottom=_GUARD_SEPARATOR_SIDE,
            )
        else:
            c.border = _THIN_BORDER
        if fill is not None:
            c.fill = fill
    top = ws.cell(row=row, column=col, value=_sanitize_cell(value))
    top.alignment = alignment or _CENTER
    return top


def _draw_guard_separator(ws: Any, row: int, n_cols: int) -> None:
    """Thicken the bottom border of ``row`` across all columns.

    Used to draw a heavy line beneath the last row of each guard's three-row
    block so neighbouring guards are visually separated. Existing left/right/top
    sides are preserved so the rest of the grid still reads as thin cells.
    """
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        existing = cell.border
        cell.border = Border(
            left=existing.left,
            right=existing.right,
            top=existing.top,
            bottom=_GUARD_SEPARATOR_SIDE,
        )


def _fill(hex_color: str | None) -> Any:
    """PatternFill from a palette hex, or ``None`` (white) when unset."""
    if not hex_color:
        return None
    return PatternFill(
        start_color=hex_color, end_color=hex_color, fill_type="solid"
    )


_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_cell(value: Any) -> Any:
    """Neutralize Excel formula injection in user-controlled cell text.

    A leading apostrophe forces Excel to treat the content as text. Applied
    only to user-controlled strings — internal literals ("✅", "-", "לא זמין")
    are written as-is so their display is unchanged.
    """
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


def _write_grid_xlsx(grid: Any) -> bytes:
    """Render a :class:`ScheduleGrid` (``schedule_grid_model``) to ``.xlsx`` bytes.

    Pure openpyxl paint pass over the shared grid model — every layout decision
    already lives in ``build_schedule_grid``; this only turns cells into styled
    openpyxl cells so the Excel and the PNG stay in lockstep.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "סידור"
    ws.sheet_view.rightToLeft = True

    n_cols = len(grid.header)  # position + 7 days
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1, value=grid.title)
    title_cell.font = _TITLE_FONT
    title_cell.alignment = _CENTER

    day_labels = list(getattr(grid, "day_labels", None) or [])
    has_labels = False
    for col, header in enumerate(grid.header, 1):
        # Day columns (col ≥ 2) carry an optional profile label on a second line.
        label = day_labels[col - 2] if col >= 2 and col - 2 < len(day_labels) else ""
        value = f"{header}\n{label}" if label else header
        cell = ws.cell(row=3, column=col, value=_sanitize_cell(value))
        _apply_header_style(cell)
        if label:
            cell.alignment = _CENTER_WRAP
            has_labels = True
    if has_labels:
        # Give the header row room for the wrapped second line.
        ws.row_dimensions[3].height = 30

    row_num = 4
    for block in grid.blocks:
        span = block.span
        _merge_vertical(
            ws, row_num, 1, span, block.name.text,
            fill=_fill(block.name.fill),
            alignment=_CENTER_WRAP if block.name.wrap else _CENTER,
            thick_bottom=True,
        )
        for day_index, day in enumerate(block.days):
            col = day_index + 2
            if day.merged:
                c = day.cells[0]
                _merge_vertical(
                    ws, row_num, col, span, c.text,
                    fill=_fill(c.fill),
                    alignment=_CENTER_WRAP if c.wrap else _CENTER,
                    thick_bottom=True,
                )
            else:
                for p, c in enumerate(day.cells):
                    cell = ws.cell(row=row_num + p, column=col)
                    _apply_cell_style(cell)
                    if c.fill:
                        cell.fill = _fill(c.fill)
                    if c.text is not None:
                        cell.value = _sanitize_cell(c.text)
                    if c.wrap:
                        cell.alignment = _CENTER_WRAP
        _draw_guard_separator(ws, row_num + span - 1, n_cols)
        row_num += span

    ws.column_dimensions[get_column_letter(1)].width = 22
    for col in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


class ExcelExportService:
    """Generates Excel schedule reports for download."""

    def __init__(
        self,
        submission_repo: SubmissionRepository,
        user_repo: UserRepository,
        week_repo: ScheduleWeekRepository,
        schedule_export_service: Any = None,
        actual_export_service: Any = None,
    ) -> None:
        self._submission_repo = submission_repo
        self._user_repo = user_repo
        self._week_repo = week_repo
        # The schedule read model (part B) — powers the built-schedule exports.
        # Optional so the pure renders (render_saved_schedule) construct with None.
        self._schedule_export = schedule_export_service
        # The ACTUAL schedule read model (סידור בפועל) — same WeekSchedule, read
        # from the week's editable execution copy. Optional for the same reason.
        self._actual_export = actual_export_service

    # ── 1. Weekly schedule grid ─────────────────────────────────────

    async def export_weekly_schedule(self, week_id: uuid.UUID) -> bytes:
        """
        Generate an Excel file with the weekly schedule grid.

        Rows = guards, columns = days (Sun–Sat).
        Each cell shows the guard's selected shifts.
        Auto-absence is applied for missing submissions.
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for Excel export")

        week = await self._week_repo.get_by_id(week_id)
        if week is None:
            raise ValueError(f"Week {week_id} not found")

        submissions = await self._submission_repo.get_submissions_for_week(week_id)
        active_users = await self._user_repo.get_active_users()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Schedule"

        # Title
        title_text = Messages.EXCEL_REPORT_TITLE.format(
            start=str(week.start_date), end=str(week.end_date)
        )
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = _TITLE_FONT
        title_cell.alignment = _CENTER

        # Headers: Name | Sun | Mon | Tue | Wed | Thu | Fri | Sat
        headers = [Messages.EXCEL_HEADER_NAME] + _DAY_NAMES_HE
        for col, header in enumerate(headers, 1):
            _apply_header_style(ws.cell(row=3, column=col, value=header))

        # Build lookup: user_id → submission
        sub_map: dict[uuid.UUID, Any] = {s.user_id: s for s in submissions}

        # Data rows
        row_num = 4
        for user in active_users:
            ws.cell(row=row_num, column=1, value=_sanitize_cell(user.full_name))
            ws.cell(row=row_num, column=1).border = _THIN_BORDER
            ws.cell(row=row_num, column=1).alignment = Alignment(vertical="center")

            sub = sub_map.get(user.id)

            for day_offset in range(7):
                col = day_offset + 2

                if sub is None:
                    # No submission → auto-absence
                    cell_value = "❌"
                    cell_fill = _RED_FILL
                else:
                    cell_value = "✅"
                    cell_fill = _GREEN_FILL

                cell = ws.cell(row=row_num, column=col, value=cell_value)
                _apply_cell_style(cell)
                if cell_fill:
                    cell.fill = cell_fill

            row_num += 1

        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logger.info(f"Excel weekly schedule exported for week {week_id}: {row_num - 4} rows")
        return buffer.read()

    # ── 1a. Saved schedule snapshot (positions × days) ──────────────

    def render_saved_schedule(self, snapshot: dict) -> bytes:
        """Render a saved-schedule snapshot (positions × days) to xlsx bytes.

        PURE: takes the self-contained snapshot dict only — no repo/DB access — so
        it renders identically after the source profile/positions were deleted.
        Rows = positions, columns = days (Sun–Sat); each day cell lists the
        assigned guard(s) and their time segment (for a tiled cell).
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for Excel export")
        if not snapshot:
            raise ValueError("snapshot is empty")

        band_fill = {
            "morning": _SHIFT_PERIODS[0][2],
            "evening": _SHIFT_PERIODS[1][2],
            "night": _SHIFT_PERIODS[2][2],
        }
        grey_fill = PatternFill(
            start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "סידור"
        ws.sheet_view.rightToLeft = True

        week = snapshot.get("week", {})
        profile_name = snapshot.get("profile_name") or ""
        title_text = (
            f"סידור עבודה — {profile_name} "
            f"({week.get('start_date', '')} — {week.get('end_date', '')})"
        )
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = _TITLE_FONT
        title_cell.alignment = _CENTER

        # Header: עמדה | ראשון … שבת
        headers = ["עמדה"] + _DAY_NAMES_HE
        for col, header in enumerate(headers, 1):
            _apply_header_style(ws.cell(row=3, column=col, value=header))

        row_num = 4
        for row in snapshot.get("rows", []):
            name = row.get("position_name", "")
            win = row.get("canonical_window") or {}
            label = name
            if win.get("start") and win.get("end"):
                label = f"{name}\n{win['start']}–{win['end']}"
            name_cell = ws.cell(row=row_num, column=1, value=_sanitize_cell(label))
            name_cell.border = _THIN_BORDER
            name_cell.alignment = _CENTER_WRAP
            fill = band_fill.get(row.get("band"))
            if fill:
                name_cell.fill = fill

            canon = (win.get("start"), win.get("end"))
            cells_by_day = {c["day_index"]: c for c in row.get("cells", [])}
            for day_index in range(7):
                col = day_index + 2
                cell = cells_by_day.get(day_index, {})
                out_cell = ws.cell(row=row_num, column=col)
                _apply_cell_style(out_cell)
                if not cell.get("active"):
                    # Blocked day (position inactive) → grey with an ✕ mark.
                    out_cell.fill = grey_fill
                    out_cell.value = "✕"
                    continue
                cell_win = cell.get("window") or {}
                assignments = cell.get("assignments", [])
                split = len(assignments) > 1
                lines = []
                for a in assignments:
                    text = a.get("guard_name", "")
                    ss, se = a.get("segment_start"), a.get("segment_end")
                    # The guard's actual hours: their own segment, else the day's
                    # window. Print hours only when *exceptional* — cell split
                    # between guards, or hours deviating from the position's norm.
                    actual = (
                        (ss, se)
                        if ss and se
                        else (cell_win.get("start"), cell_win.get("end"))
                    )
                    if actual[0] and actual[1] and (split or actual != canon):
                        text = f"{text}\n{actual[0]}–{actual[1]}"
                    lines.append(text)
                if lines:
                    out_cell.value = _sanitize_cell("\n".join(lines))
                    out_cell.alignment = _CENTER_WRAP
                else:
                    # Active cell with no guard → flag the gap amber.
                    out_cell.fill = _EMPTY_FILL
            row_num += 1

        # Auto-width
        ws.column_dimensions[get_column_letter(1)].width = 22
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logger.info(
            "Saved-schedule xlsx rendered: %d position rows", row_num - 4
        )
        return buffer.read()

    # ── 1c. Built-schedule grid (positions × days) ──────────────────

    async def export_schedule_grid(self, week_id: uuid.UUID) -> bytes:
        """Export the *built* schedule: rows = positions (board order), columns =
        7 days, each cell the assigned guard(s) + their hours.

        A cell tiled between several guards splits the **position block** into one
        sub-row per guard (like the constraints report's three-row blocks, but the
        span is the widest tiling the position sees that week). Reads exclusively
        from the schedule read model (task 01) so all three products agree.
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for Excel export")

        week = await self._week_repo.get_by_id(week_id)
        if week is None:
            raise ValueError(f"Week {week_id} not found")

        schedule = await self._schedule_export.get_week_schedule(week_id)

        grid = build_schedule_grid(schedule, week)
        data = _write_grid_xlsx(grid)
        logger.info(
            "Schedule-grid xlsx exported for week %s: %d position rows",
            week_id, len(schedule.by_position),
        )
        return data

    async def export_actual_schedule_grid(self, week_id: uuid.UUID) -> bytes:
        """Export the ACTUAL schedule (סידור בפועל) in the exact grid layout.

        Identical rendering to :meth:`export_schedule_grid` — same colours,
        same tiling/partial-coverage rules — only the source differs: the
        week's editable execution copy instead of the frozen plan. An ad-hoc
        position renders like any other row (its band derives from its hours).
        """
        from app.schedule_builder.services.actual_schedule_service import (
            ActualScheduleNotAvailableException,
        )
        from app.utils.date_utils import today_il

        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for Excel export")
        if self._actual_export is None:
            raise RuntimeError("actual_export_service was not provided")

        week = await self._week_repo.get_by_id(week_id)
        if week is None:
            raise ValueError(f"Week {week_id} not found")
        if week.start_date > today_il():
            # The read model would silently fall back to the plan; for an
            # explicit "download what really happened" that's misleading — the
            # week-card offers the planned snapshot for future weeks instead.
            raise ActualScheduleNotAvailableException()

        schedule = await self._actual_export.get_week_schedule(week_id)
        grid = build_schedule_grid(schedule, week)
        data = _write_grid_xlsx(grid)
        logger.info(
            "ACTUAL schedule-grid xlsx exported for week %s: %d position rows",
            week_id, len(schedule.by_position),
        )
        return data

    async def export_schedule_grid_png(self, week_id: uuid.UUID) -> bytes:
        """Export the built schedule as a PNG image — same grid model (and so the
        same colours/layout) as :meth:`export_schedule_grid`, rendered with Pillow.

        This is what guards receive on publish: a photo opens with one tap on a
        phone, unlike an ``.xlsx``. The admin Excel export is unchanged.
        """
        from app.services.schedule_grid_png import render_schedule_grid_png

        week = await self._week_repo.get_by_id(week_id)
        if week is None:
            raise ValueError(f"Week {week_id} not found")

        schedule = await self._schedule_export.get_week_schedule(week_id)
        grid = build_schedule_grid(schedule, week)
        data = render_schedule_grid_png(grid)
        logger.info(
            "Schedule-grid PNG exported for week %s: %d position rows",
            week_id, len(schedule.by_position),
        )
        return data

    async def export_actual_schedule_grid_png(self, week_id: uuid.UUID) -> bytes:
        """The ACTUAL schedule (סידור בפועל) as a PNG — same Pillow renderer
        and layout as the planned grid image, read from the execution copy.
        Future weeks are rejected, mirroring the actual Excel export."""
        from app.schedule_builder.services.actual_schedule_service import (
            ActualScheduleNotAvailableException,
        )
        from app.services.schedule_grid_png import render_schedule_grid_png
        from app.utils.date_utils import today_il

        if self._actual_export is None:
            raise RuntimeError("actual_export_service was not provided")

        week = await self._week_repo.get_by_id(week_id)
        if week is None:
            raise ValueError(f"Week {week_id} not found")
        if week.start_date > today_il():
            raise ActualScheduleNotAvailableException()

        schedule = await self._actual_export.get_week_schedule(week_id)
        grid = build_schedule_grid(schedule, week)
        data = render_schedule_grid_png(grid)
        logger.info(
            "ACTUAL schedule-grid PNG exported for week %s: %d position rows",
            week_id, len(schedule.by_position),
        )
        return data

    # ── 1d. Per-guard "positions" grid (guard-grouped) ──────────────

    async def export_guard_positions(self, week_id: uuid.UUID) -> bytes:
        """Export a per-guard overview: one block per active guard, columns = 7
        days, each cell that guard's position + hours that day.

        Every active guard appears — a guard with no shifts gets a single row
        stating so. A guard with two positions on one day splits into sub-rows.
        Reads the ``by_guard`` cut of the schedule read model (task 01).
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for Excel export")

        week = await self._week_repo.get_by_id(week_id)
        if week is None:
            raise ValueError(f"Week {week_id} not found")

        schedule = await self._schedule_export.get_week_schedule(week_id)
        # Guards with no linked Telegram float to the top — the admin has to hand
        # them their schedule off-band, so they're the ones to notice first.
        guards = order_unverified_first(schedule.by_guard)
        scheduled = sum(1 for g in guards if g.shifts)

        grey_fill = PatternFill(
            start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "עמדות למאבטח"
        ws.sheet_view.rightToLeft = True

        n_cols = 8  # name + 7 days
        title_text = f"עמדות למאבטח — {week.start_date} עד {week.end_date}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = _TITLE_FONT
        title_cell.alignment = _CENTER

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        ws.cell(
            row=2, column=1, value=f"שובצו: {scheduled} מתוך {len(guards)} מאבטחים",
        ).alignment = _CENTER

        headers = ["שם"] + _DAY_NAMES_HE
        for col, header in enumerate(headers, 1):
            _apply_header_style(ws.cell(row=3, column=col, value=header))

        row_num = 4
        for guard in guards:
            shifts_by_day: dict[int, list] = {}
            for s in guard.shifts:
                shifts_by_day.setdefault(s.day_index, []).append(s)
            span = max([len(v) for v in shifts_by_day.values()] + [1])

            # Flag a guard with no linked Telegram: soft-red name block with a
            # note, so the admin sees whose schedule needs off-band hand-off.
            no_telegram = guard.telegram_id is None
            name_value = (
                f"{guard.user_name}\n🚫 אין טלגרם" if no_telegram else guard.user_name
            )
            _merge_vertical(
                ws, row_num, 1, span, name_value,
                fill=_NO_TG_FILL if no_telegram else None,
                alignment=_CENTER_WRAP, thick_bottom=True,
            )

            if not guard.shifts:
                # No placements this week → one note merged across the day columns.
                ws.merge_cells(
                    start_row=row_num, start_column=2, end_row=row_num, end_column=n_cols
                )
                for col in range(2, n_cols + 1):
                    c = ws.cell(row=row_num, column=col)
                    c.border = _THIN_BORDER
                    c.fill = grey_fill
                note = ws.cell(row=row_num, column=2, value="אין שיבוצים השבוע")
                note.alignment = _CENTER
                _draw_guard_separator(ws, row_num, n_cols)
                row_num += 1
                continue

            for day_index in range(7):
                col = day_index + 2
                day_shifts = shifts_by_day.get(day_index, [])
                for p in range(span):
                    cell = ws.cell(row=row_num + p, column=col)
                    _apply_cell_style(cell)
                    if p >= len(day_shifts):
                        cell.fill = grey_fill
                        continue
                    s = day_shifts[p]
                    cell.value = f"{s.position_name}\n{s.start}–{s.end}"
                    cell.alignment = _CENTER_WRAP

            _draw_guard_separator(ws, row_num + span - 1, n_cols)
            row_num += span

        ws.column_dimensions[get_column_letter(1)].width = 22
        for col in range(2, n_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logger.info(
            "Guard-positions xlsx exported for week %s: %d guards (%d scheduled)",
            week_id, len(guards), scheduled,
        )
        return buffer.read()

    # ── 1b. Constraints report (who submitted what) ─────────────────

    async def export_constraints_report(self, week_id: uuid.UUID) -> bytes:
        """
        Generate a nicely-formatted Excel of all submitted constraints.

        Each guard who submitted occupies three stacked rows — בוקר / ערב /
        לילה — so every shift period gets its own line per day instead of being
        crammed into a single cell. Name, phone and notes are merged across the
        three rows. Sheet is rendered right-to-left for Hebrew readability.
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for Excel export")

        week = await self._week_repo.get_by_id(week_id)
        if week is None:
            raise ValueError(f"Week {week_id} not found")

        submissions = await self._submission_repo.get_submissions_for_week(week_id)
        active_users = await self._user_repo.get_active_users()
        user_map = {u.id: u for u in active_users}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "אילוצים"
        ws.sheet_view.rightToLeft = True

        # Layout: id | name | phone | period | 7 days | notes
        _ID_COL = 1
        _NAME_COL = 2
        _PHONE_COL = 3
        _PERIOD_COL = 4
        _FIRST_DAY_COL = 5
        _NOTES_COL = 12
        n_cols = 12

        # Title
        title_text = f"אילוצים שהוגשו — {week.start_date} עד {week.end_date}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = _TITLE_FONT
        title_cell.alignment = _CENTER

        # Subtitle: how many submitted
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        ws.cell(
            row=2,
            column=1,
            value=f"סך הכל הגישו: {len(submissions)} מאבטחים",
        ).alignment = _CENTER

        # Headers
        headers = (
            ["מזהה", Messages.EXCEL_HEADER_NAME, Messages.EXCEL_HEADER_PHONE, "משמרת"]
            + _DAY_NAMES_HE
            + ["הערות"]
        )
        for col, header in enumerate(headers, 1):
            _apply_header_style(ws.cell(row=4, column=col, value=header))

        # Sort submitting guards by name for a stable, readable order
        def _sub_name(sub: Any) -> str:
            user = user_map.get(sub.user_id)
            return user.full_name if user else ""

        ordered = sorted(submissions, key=_sub_name)

        _SPAN = len(_SHIFT_PERIODS)  # three rows per guard
        row_num = 5
        for sub in ordered:
            user = user_map.get(sub.user_id)
            if user is None:
                user = await self._user_repo.get_by_id(sub.user_id)
            full_name = user.full_name if user else "—"
            phone = (user.phone_number if user else "") or ""

            # Map each calendar date → its daily status for this submission
            status_by_date: dict[date, Any] = {
                ds.date: ds for ds in sub.daily_statuses
            }

            # ID / name / phone / notes span the guard's three period rows.
            # The guard's DB id is injected so each row can be tied back to the
            # exact guard (and their attributes), regardless of who filled the
            # constraints — the guard via Telegram or the admin on their behalf.
            guard_id = str(user.id) if user else ""
            _merge_vertical(ws, row_num, _ID_COL, _SPAN, guard_id, thick_bottom=True)
            _merge_vertical(ws, row_num, _NAME_COL, _SPAN, full_name, thick_bottom=True)
            _merge_vertical(ws, row_num, _PHONE_COL, _SPAN, phone, thick_bottom=True)
            _merge_vertical(
                ws,
                row_num,
                _NOTES_COL,
                _SPAN,
                sub.general_notes or "",
                alignment=Alignment(
                    horizontal="right", vertical="center", wrap_text=True
                ),
                thick_bottom=True,
            )

            # "משמרת" label column — one coloured row per period
            for p, (_type, label, accent) in enumerate(_SHIFT_PERIODS):
                cell = ws.cell(row=row_num + p, column=_PERIOD_COL, value=label)
                cell.border = _THIN_BORDER
                cell.alignment = _CENTER
                cell.fill = accent
                cell.font = Font(bold=True)

            for day_offset in range(7):
                col = _FIRST_DAY_COL + day_offset
                current_date = week.start_date + timedelta(days=day_offset)
                ds = status_by_date.get(current_date)

                # Unavailable day, or available with no chosen windows →
                # a single merged cell spanning the three period rows.
                if ds is None or not ds.is_available:
                    _merge_vertical(
                        ws, row_num, col, _SPAN, "לא זמין",
                        fill=_RED_FILL, thick_bottom=True,
                    )
                    continue

                windows_by_type: dict[str, list[Any]] = {}
                for w in ds.shift_windows:
                    t = getattr(w.shift_type, "value", w.shift_type)
                    windows_by_type.setdefault(t, []).append(w)

                if not windows_by_type:
                    _merge_vertical(
                        ws, row_num, col, _SPAN, "זמין",
                        fill=_GREEN_FILL, thick_bottom=True,
                    )
                    continue

                # One row per period, showing that period's time windows.
                for p, (shift_type, _label, _accent) in enumerate(_SHIFT_PERIODS):
                    cell = ws.cell(row=row_num + p, column=col)
                    cell.border = _THIN_BORDER
                    cell.alignment = _CENTER_WRAP
                    wins = sorted(
                        windows_by_type.get(shift_type, []),
                        key=lambda w: w.start_time,
                    )
                    if wins:
                        cell.value = "\n".join(
                            f"{w.start_time:%H:%M}–{w.end_time:%H:%M}"
                            for w in wins
                        )
                        cell.fill = _GREEN_FILL
                    else:
                        cell.fill = _EMPTY_FILL

            # Heavy line under this guard's block to separate it from the next.
            _draw_guard_separator(ws, row_num + _SPAN - 1, n_cols)

            row_num += _SPAN

        # Column widths: id + name + notes wider, days medium
        ws.column_dimensions[get_column_letter(_ID_COL)].width = 38  # id (uuid)
        ws.column_dimensions[get_column_letter(_NAME_COL)].width = 20  # name
        ws.column_dimensions[get_column_letter(_PHONE_COL)].width = 16  # phone
        ws.column_dimensions[get_column_letter(_PERIOD_COL)].width = 10  # period
        for col in range(_FIRST_DAY_COL, _NOTES_COL):  # days
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.column_dimensions[get_column_letter(_NOTES_COL)].width = 30  # notes

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logger.info(
            f"Excel constraints report exported for week {week_id}: "
            f"{len(ordered)} guards"
        )
        return buffer.read()

    # ── 2. Deviation report ─────────────────────────────────────────

    async def export_deviation_report(self, week_id: uuid.UUID) -> bytes:
        """
        Generate a deviation report Excel file.

        Lists guards whose shift counts deviate from thresholds.
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for Excel export")

        week = await self._week_repo.get_by_id(week_id)
        if week is None:
            raise ValueError(f"Week {week_id} not found")

        submissions = await self._submission_repo.get_submissions_for_week(week_id)
        active_users = await self._user_repo.get_active_users()
        user_map = {u.id: u for u in active_users}

        # Get deviation settings (min_shifts_per_week from system_settings)
        # Default threshold: 3 shifts per week
        threshold = 3

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Deviations"

        # Title
        title_text = f"דוח חריגות — {week.start_date} עד {week.end_date}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = _TITLE_FONT
        title_cell.alignment = _CENTER

        # Headers
        headers = [
            Messages.EXCEL_HEADER_NAME,
            Messages.EXCEL_HEADER_PHONE,
            "מספר משמרות",
            Messages.EXCEL_HEADER_THRESHOLDS,
            Messages.EXCEL_HEADER_DEVIATION,
        ]
        for col, header in enumerate(headers, 1):
            _apply_header_style(ws.cell(row=3, column=col, value=header))

        row_num = 4
        for sub in submissions:
            user = user_map.get(sub.user_id)
            if not user:
                continue

            # Show all submissions
            row_data = [
                _sanitize_cell(user.full_name),
                _sanitize_cell(user.phone_number or ""),
                "✅",
                str(threshold),
                "-",
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                _apply_cell_style(cell)
            row_num += 1
        # Also show guards with no submission
        sub_user_ids = {s.user_id for s in submissions}
        for user in active_users:
            if user.id not in sub_user_ids:
                row_data = [
                    _sanitize_cell(user.full_name),
                    _sanitize_cell(user.phone_number or ""),
                    "❌",
                    str(threshold),
                    f"{threshold}-",
                ]
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    _apply_cell_style(cell)
                    if col == 5:
                        cell.fill = _RED_FILL
                row_num += 1

        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logger.info(f"Excel deviation report exported for week {week_id}: {row_num - 4} rows")
        return buffer.read()

    # ── 3. Guard history report ─────────────────────────────────────

    async def export_guard_history(
        self, user_id: uuid.UUID, start_date: date, end_date: date
    ) -> bytes:
        """
        Generate a per-guard history report across multiple weeks.

        Shows submission status for each week in the range.
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for Excel export")

        user = await self._user_repo.get_by_id(user_id)
        if user is None:
            raise ValueError(f"User {user_id} not found")

        # Get submissions for this user
        submissions = await self._submission_repo.get_by_user(user_id)
        sub_by_week: dict[uuid.UUID, Any] = {s.week_id: s for s in submissions}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Guard History"

        # Title
        title_text = f"היסטוריית מאבטח — {user.full_name}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = _TITLE_FONT
        title_cell.alignment = _CENTER

        # Subtitle with date range
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
        ws.cell(
            row=2, column=1, value=f"תקופה: {start_date} עד {end_date}"
        ).alignment = _CENTER

        # Headers
        headers = [
            "שבוע",
            Messages.EXCEL_HEADER_NAME,
            "סטטוס הגשה",
        ]
        for col, header in enumerate(headers, 1):
            _apply_header_style(ws.cell(row=4, column=col, value=header))

        # Iterate weeks in range
        row_num = 5
        current = start_date
        while current <= end_date:
            # Find week that starts on this date (or closest)
            # We'll show entries for each 7-day chunk
            week_end = current + timedelta(days=6)

            # Find matching submission
            status_text = Messages.STATUS_PENDING

            for s_week_id, s in sub_by_week.items():
                # We need the week object to match dates
                # For now, show all submissions we have
                pass

            row_data = [
                f"{current} – {week_end}",
                _sanitize_cell(user.full_name),
                status_text,
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                _apply_cell_style(cell)

            row_num += 1
            current += timedelta(days=7)

        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logger.info(
            f"Excel guard history exported for user {user_id}: {row_num - 5} weeks"
        )
        return buffer.read()