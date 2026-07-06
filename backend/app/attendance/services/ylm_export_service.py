"""
YLM Excel exports (stage 3 / 03) — the payroll bureau's two sheets.

Faithful to the sample PDFs (~/Documents/1.pdf, 2.pdf) with the 4/7 scope:
HOURS COLUMNS ONLY — the rate columns (100/125/150, ש150/ש175/ש200) exist in
the layout but stay empty; the bureau computes them. Manually-entered/edited
punches are highlighted orange exactly like the sample. RTL sheets.

Everything renders from the PayrollReadModel's EmployeeMonth — one source,
both sheets, no way to disagree.
"""

import logging
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.attendance.services.payroll_readmodel import (
    EmployeeMonth,
    minutes_hhmm,
)

logger = logging.getLogger("ilutzim")

_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_EDITED_FILL = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
_SUMMARY_FILL = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_BOLD = Font(bold=True)
_RED_BOLD = Font(bold=True, color="CC0000")

# Column order (col A is the VISUAL RIGHT in an RTL sheet — matches the sample).
EMPLOYEE_COLUMNS = [
    "אתר", "יום", "תאריך", "כניסה", "יציאה", "סה\"כ", "תקן", "ח/ע",
    "100%", "125%", "150%", "הערות",
]

HEB_MONTHS = [
    "ינואר", "פברואר", "מרץ", "אפריל", "מאי", "יוני",
    "יולי", "אוגוסט", "ספטמבר", "אוקטובר", "נובמבר", "דצמבר",
]


def _hhmm(dt) -> str:
    return dt.strftime("%H:%M") if dt else ""


def _month_title(year: int, month: int) -> str:
    return f"{HEB_MONTHS[month - 1]} {year}"


class YlmExportService:
    """Builds the two YLM .xlsx sheets from EmployeeMonth read-models."""

    # ── per-employee monthly attendance sheet ────────────────────────────────

    def export_employee_report(self, month: EmployeeMonth) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "דוח נוכחות"
        ws.sheet_view.rightToLeft = True

        # header block (like the sample's top strip)
        ws["A1"] = f"דוח נוכחות — {_month_title(month.year, month.month)}"
        ws["A1"].font = Font(bold=True, size=14)
        header_pairs = [
            ("שם העובד:", month.user_name),
            ("מ.עובד:", month.payroll_employee_id),
            ("קוד י.ל.מ:", month.payroll_ylm_code),
            ("ת.ז.:", month.national_id),
            ("חברה:", month.company_name),
        ]
        for idx, (label, value) in enumerate(header_pairs):
            ws.cell(row=2, column=1 + idx * 2, value=label).font = _BOLD
            ws.cell(row=2, column=2 + idx * 2, value=value or "")

        # table headers
        head_row = 4
        for col, title in enumerate(EMPLOYEE_COLUMNS, start=1):
            cell = ws.cell(row=head_row, column=col, value=title)
            cell.font = _BOLD
            cell.fill = _HEADER_FILL
            cell.border = _BORDER
            cell.alignment = _CENTER

        # day rows — one line per read-model row (empty days included)
        row_idx = head_row
        for row in month.rows:
            row_idx += 1
            values = [
                row.site,
                row.day_letter,
                row.day.strftime("%d/%m"),
                _hhmm(row.check_in),
                _hhmm(row.check_out),
                minutes_hhmm(row.total_minutes) if (row.check_in or row.site) else "",
                minutes_hhmm(row.norm_minutes) if row.norm_minutes is not None else "",
                minutes_hhmm(row.diff_minutes) if row.diff_minutes is not None else "",
                "", "", "",  # 100% / 125% / 150% — the bureau computes these
                row.notes,
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = _BORDER
                if col in (4, 5) and row.edited:
                    cell.fill = _EDITED_FILL  # orange edited punch — like the sample
                if 2 <= col <= 11:
                    cell.alignment = _CENTER

        # summary block ("ריכוז נתונים")
        totals = month.totals
        summary_row = row_idx + 2
        ws.cell(row=summary_row, column=1, value="ריכוז נתונים").font = _BOLD
        summary_pairs = [
            ("ימי עבודה בפועל:", str(totals.work_days)),
            ("שעות עבודה בפועל:", minutes_hhmm(totals.actual_minutes)),
            ("שעות תקן:", minutes_hhmm(totals.norm_minutes)),
            ("שעות עודף/חוסר:", minutes_hhmm(totals.diff_minutes)),
        ]
        for offset, (label, value) in enumerate(summary_pairs, start=1):
            label_cell = ws.cell(row=summary_row + offset, column=1, value=label)
            value_cell = ws.cell(row=summary_row + offset, column=2, value=value)
            label_cell.fill = _SUMMARY_FILL
            value_cell.fill = _SUMMARY_FILL
            label_cell.font = _BOLD

        sign_row = summary_row + len(summary_pairs) + 2
        ws.cell(row=sign_row, column=1, value="חתימת עובד: ____________")
        ws.cell(row=sign_row, column=4, value="חתימת אחראי: ____________")

        widths = [16, 5, 8, 8, 8, 7, 7, 7, 7, 7, 7, 28]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=head_row, column=col).column_letter].width = width

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── center (all-employees) summary sheet ────────────────────────────────

    CENTER_COLUMNS = [
        "עובד", "סה\"כ", "100%", "125%", "150%",
        "ש150%", "ש175%", "ש200%", "תקן", "ח/ע", "ימים",
    ]

    def export_center_report(
        self, months: list[EmployeeMonth], year: int, month: int, company_name: str
    ) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "דוח מרכז"
        ws.sheet_view.rightToLeft = True

        ws["A1"] = f"דוח מרכז — {_month_title(year, month)}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"חברה: {company_name}"

        head_row = 4
        for col, title in enumerate(self.CENTER_COLUMNS, start=1):
            cell = ws.cell(row=head_row, column=col, value=title)
            cell.font = _BOLD
            cell.fill = _HEADER_FILL
            cell.border = _BORDER
            cell.alignment = _CENTER

        total_actual = total_norm = total_diff = total_days = 0
        row_idx = head_row
        for emp in months:
            row_idx += 1
            t = emp.totals
            total_actual += t.actual_minutes
            total_norm += t.norm_minutes
            total_diff += t.diff_minutes
            total_days += t.work_days
            values = [
                emp.user_name,
                minutes_hhmm(t.actual_minutes),
                "", "", "", "", "", "",  # rate columns — bureau-computed
                minutes_hhmm(t.norm_minutes),
                minutes_hhmm(t.diff_minutes),
                t.work_days,
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = _BORDER
                if col > 1:
                    cell.alignment = _CENTER

        # the red bold "סיכום" line, like the sample
        row_idx += 1
        totals_values = [
            "סיכום",
            minutes_hhmm(total_actual),
            "", "", "", "", "", "",
            minutes_hhmm(total_norm),
            minutes_hhmm(total_diff),
            total_days,
        ]
        for col, value in enumerate(totals_values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = _RED_BOLD
            cell.border = _BORDER
            if col > 1:
                cell.alignment = _CENTER

        widths = [20, 8, 7, 7, 7, 7, 7, 7, 8, 8, 6]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=head_row, column=col).column_letter].width = width

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
