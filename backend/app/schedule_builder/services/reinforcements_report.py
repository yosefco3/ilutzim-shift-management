"""
Reinforcements report (step 13) — "דוח מתגברים".

The external helpers don't punch and never reach payroll, so their engagement
is settled off-system — this report is the paper trail: who reinforced, on
which date, for which hours. The caller picks the period (the UI offers
daily / weekly / monthly cuts; here it is just [start, end]).

Hours come from the actual board (segment when set, else the position's day
window) — the same numbers the actual schedule shows. A shift crossing
midnight (end <= start) counts past 24:00, attributed to its start date.
"""

import io
import logging
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.schedule_builder.utils import intervals as iv

logger = logging.getLogger("ilutzim")

_DAY_LETTERS = "אבגדהוש"

_HEAD_FILL = PatternFill("solid", fgColor="1F2937")
_HEAD_FONT = Font(bold=True, color="FFFFFF")
_TOTAL_FILL = PatternFill("solid", fgColor="FDE68A")
_THIN = Side(style="thin", color="D1D5DB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")

_COLUMNS = ["תאריך", "יום", "שם המתגבר", "טלפון", "מפקח", "עמדה", "שעות", "משך"]
_WIDTHS = [12, 6, 20, 14, 16, 18, 14, 9]


def _hhmm_total(minutes: int) -> str:
    return f"{minutes // 60}:{minutes % 60:02d}"


def _display_phone(phone: str | None) -> str:
    # The auto-generated placeholder is an internal detail, not a phone.
    if not phone or phone.startswith("EXT-"):
        return "—"
    return phone


def collect_report_rows(assignments, cards_by_key, start: date, end: date) -> list[dict]:
    """Resolve loaded assignments into dated report rows within [start, end]."""
    rows: list[dict] = []
    for a in assignments:
        week = a.actual_schedule.week
        work_date = week.start_date + timedelta(days=a.day_index)
        if not (start <= work_date <= end):
            continue
        window = (a.actual_position.day_schedules or {}).get(str(a.day_index))
        if a.segment_start and a.segment_end:
            hours = (a.segment_start, a.segment_end)
        elif window:
            hours = (window["start"], window["end"])
        else:  # assignment stranded on a deactivated day — nothing to bill
            continue
        card = cards_by_key.get((a.user_id, a.actual_schedule_id))
        rows.append({
            "date": work_date,
            "day_letter": _DAY_LETTERS[a.day_index],
            "name": a.user.full_name,
            "phone": _display_phone(a.user.phone_number),
            "supervisor": (card.supervisor_name if card else None) or "—",
            "position": a.actual_position.name,
            "start": hours[0],
            "end": hours[1],
            "minutes": iv.duration(iv.normalize(*hours)),
        })
    rows.sort(key=lambda r: (r["name"], r["date"], r["start"]))
    return rows


def render_reinforcements_report(rows: list[dict], start: date, end: date) -> bytes:
    """Rows → the RTL xlsx: per-guard blocks with subtotals + a grand total."""
    wb = Workbook()
    ws = wb.active
    ws.title = "דוח מתגברים"
    ws.sheet_view.rightToLeft = True
    for i, width in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = width

    ws.append([f"דוח מתגברים — {start.strftime('%d/%m/%Y')} עד {end.strftime('%d/%m/%Y')}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_COLUMNS))
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.append([])

    ws.append(_COLUMNS)
    header_row = ws.max_row
    for col in range(1, len(_COLUMNS) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = _HEAD_FILL
        cell.font = _HEAD_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER

    grand_total = 0
    current_name = None
    block_total = 0

    def _close_block():
        nonlocal block_total
        if current_name is None:
            return
        ws.append([
            "", "", f'סה"כ {current_name}', "", "", "", "",
            _hhmm_total(block_total),
        ])
        for col in range(1, len(_COLUMNS) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill = _TOTAL_FILL
            cell.font = Font(bold=True)
        block_total = 0

    for row in rows:
        if row["name"] != current_name:
            _close_block()
            current_name = row["name"]
        ws.append([
            row["date"].strftime("%d/%m/%Y"),
            row["day_letter"],
            row["name"],
            row["phone"],
            row["supervisor"],
            row["position"],
            f'{row["start"]}–{row["end"]}',
            _hhmm_total(row["minutes"]),
        ])
        for col in range(1, len(_COLUMNS) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.border = _BORDER
            cell.alignment = _CENTER
        block_total += row["minutes"]
        grand_total += row["minutes"]
    _close_block()

    ws.append([])
    ws.append(["", "", 'סה"כ כללי', "", "", "", "", _hhmm_total(grand_total)])
    for col in range(1, len(_COLUMNS) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = Font(bold=True, size=12)

    if not rows:
        ws.append([])
        ws.append(["אין שיבוצי מתגברים בתקופה זו"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
