"""
Step 07 — the actual-schedule Excel export ("what really happened").
"""

import io
from datetime import timedelta

import pytest
from openpyxl import load_workbook

from app.repositories.schedule_week_repository import ScheduleWeekRepository
from app.repositories.submission_repository import SubmissionRepository
from app.repositories.user_repository import UserRepository
from app.schedule_builder.services.actual_schedule_service import (
    ActualScheduleNotAvailableException,
)
from app.services.excel_export_service import ExcelExportService

from tests.test_actual_export_parity import (
    _actual_export,
    _actual_service,
    _make_rich_week,
    _planned_export,
)


def _excel_service(session) -> ExcelExportService:
    return ExcelExportService(
        SubmissionRepository(session),
        UserRepository(session),
        ScheduleWeekRepository(session),
        _planned_export(session),
        _actual_export(session),
    )


def _all_texts(data: bytes) -> str:
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    return " ".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    )


@pytest.mark.asyncio
async def test_actual_export_reflects_mid_week_edits(db_session):
    week = await _make_rich_week(db_session)

    # Edit the actual copy: add an ad-hoc position with a guard on it.
    service = _actual_service(db_session)
    added = await service.add_position(
        week.id, name="אבטחת אירוע חד פעמי",
        day_schedules={"5": {"start": "16:00", "end": "20:00"}},
    )
    from app.models.user import User

    substitute = User(
        phone_number="0509999999", first_name="גיא", last_name="מחליף",
    )
    db_session.add(substitute)
    await db_session.commit()
    await service.assign(added.id, 5, substitute.id)

    data = await _excel_service(db_session).export_actual_schedule_grid(week.id)
    assert data[:2] == b"PK"  # a real xlsx (zip container)
    texts = _all_texts(data)
    assert "אבטחת אירוע חד פעמי" in texts
    assert "גיא מחליף" in texts

    # The PLANNED export stays untouched by the actual edit.
    planned_data = await _excel_service(db_session).export_schedule_grid(week.id)
    assert "אבטחת אירוע חד פעמי" not in _all_texts(planned_data)


@pytest.mark.asyncio
async def test_actual_export_rejects_a_future_week(db_session):
    week = await _make_rich_week(db_session, started=False)

    with pytest.raises(ActualScheduleNotAvailableException):
        await _excel_service(db_session).export_actual_schedule_grid(week.id)


@pytest.mark.asyncio
async def test_actual_png_export_renders_image(db_session):
    """The PNG twin of the actual Excel — real image bytes, same source."""
    week = await _make_rich_week(db_session)
    data = await _excel_service(db_session).export_actual_schedule_grid_png(week.id)
    assert data[:8] == b"\x89PNG\r\n\x1a\n"


@pytest.mark.asyncio
async def test_actual_png_export_rejects_a_future_week(db_session):
    week = await _make_rich_week(db_session, started=False)
    with pytest.raises(ActualScheduleNotAvailableException):
        await _excel_service(db_session).export_actual_schedule_grid_png(week.id)
