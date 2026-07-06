"""
Stage 3 / 03 steps 2+3 — the two YLM xlsx sheets, rendered from EmployeeMonth.
"""

import uuid
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook

from app.attendance.services.payroll_readmodel import (
    EmployeeMonth,
    MonthRow,
    MonthTotals,
)
from app.attendance.services.ylm_export_service import (
    EMPLOYEE_COLUMNS,
    YlmExportService,
)


def _month(rows=None, **overrides) -> EmployeeMonth:
    base = dict(
        user_id=uuid.uuid4(),
        user_name="יוסי כהן",
        national_id="034465773",
        payroll_employee_id="605182",
        payroll_ylm_code="605182",
        company_name="ספרא",
        year=2026,
        month=7,
        rows=rows or [],
        totals=MonthTotals(
            work_days=21, actual_minutes=10778, norm_minutes=12795,
            diff_minutes=-2017,
        ),
    )
    base.update(overrides)
    return EmployeeMonth(**base)


def _row(**overrides) -> MonthRow:
    base = dict(
        day=date(2026, 7, 5),
        day_letter="א",
        site="שער ראשי",
        check_in=datetime(2026, 7, 5, 7, 2),
        check_out=datetime(2026, 7, 5, 15, 15),
        check_out_raw=datetime(2026, 7, 5, 15, 1),
        total_minutes=493,
        norm_minutes=480,
        diff_minutes=13,
        notes="יציאה בפועל 15:01",
        edited=False,
    )
    base.update(overrides)
    return MonthRow(**base)


def _sheet(data: bytes):
    return load_workbook(BytesIO(data)).active


def test_employee_sheet_layout_and_values():
    data = YlmExportService().export_employee_report(
        _month(rows=[
            _row(),
            _row(day=date(2026, 7, 6), day_letter="ב", site="", check_in=None,
                 check_out=None, check_out_raw=None, total_minutes=0,
                 norm_minutes=None, diff_minutes=None, notes=""),
        ])
    )
    ws = _sheet(data)

    assert ws.sheet_view.rightToLeft is True
    assert "יולי 2026" in ws["A1"].value
    header_values = {ws.cell(row=2, column=c).value for c in range(1, 11)}
    assert {"שם העובד:", "יוסי כהן", "מ.עובד:", "605182", "ת.ז.:",
            "034465773", "חברה:", "ספרא"} <= header_values

    # column headers
    for col, title in enumerate(EMPLOYEE_COLUMNS, start=1):
        assert ws.cell(row=4, column=col).value == title

    # the worked row: exact in, rounded out, totals, empty rate columns, note
    assert ws.cell(row=5, column=1).value == "שער ראשי"
    assert ws.cell(row=5, column=2).value == "א"
    assert ws.cell(row=5, column=3).value == "05/07"
    assert ws.cell(row=5, column=4).value == "07:02"
    assert ws.cell(row=5, column=5).value == "15:15"
    assert ws.cell(row=5, column=6).value == "8:13"
    assert ws.cell(row=5, column=7).value == "8:00"
    assert ws.cell(row=5, column=8).value == "0:13"
    assert ws.cell(row=5, column=9).value in ("", None)   # 100% — bureau's job
    assert ws.cell(row=5, column=11).value in ("", None)  # 150%
    assert "יציאה בפועל 15:01" in ws.cell(row=5, column=12).value

    # the empty calendar day still prints its letter+date (times stay blank)
    assert ws.cell(row=6, column=2).value == "ב"
    assert ws.cell(row=6, column=4).value in ("", None)

    # summary block with the negative diff
    texts = [
        (c.value, ws.cell(row=c.row, column=2).value)
        for row_cells in ws.iter_rows(min_row=7, max_row=20, max_col=1)
        for c in row_cells if c.value
    ]
    flat = dict(t for t in texts if t[0])
    assert flat.get("ימי עבודה בפועל:") == "21"
    assert flat.get("שעות עודף/חוסר:") == "-33:37"
    assert any("חתימת עובד" in (c.value or "") for row in ws.iter_rows(max_col=1) for c in row)


def test_employee_sheet_marks_edited_punches_orange():
    data = YlmExportService().export_employee_report(
        _month(rows=[_row(edited=True, notes="הוזן/תוקן ידנית")])
    )
    ws = _sheet(data)
    assert ws.cell(row=5, column=4).fill.start_color.rgb.endswith("FCD5B4")
    assert ws.cell(row=5, column=5).fill.start_color.rgb.endswith("FCD5B4")
    # non-punch cells stay unfilled
    assert not (ws.cell(row=5, column=6).fill.start_color.rgb or "").endswith("FCD5B4")


def test_center_sheet_rows_and_red_total():
    months = [
        _month(user_name="אבי אלף",
               totals=MonthTotals(work_days=20, actual_minutes=9600,
                                  norm_minutes=9600, diff_minutes=0)),
        _month(user_name="יוסי כהן",
               totals=MonthTotals(work_days=21, actual_minutes=10778,
                                  norm_minutes=12795, diff_minutes=-2017)),
    ]
    data = YlmExportService().export_center_report(months, 2026, 7, "ספרא")
    ws = _sheet(data)

    assert ws.sheet_view.rightToLeft is True
    assert ws.cell(row=4, column=1).value == "עובד"
    assert ws.cell(row=5, column=1).value == "אבי אלף"
    assert ws.cell(row=5, column=2).value == "160:00"
    assert ws.cell(row=6, column=10).value == "-33:37"
    # the summary line, bold red, sums the filled columns
    assert ws.cell(row=7, column=1).value == "סיכום"
    assert ws.cell(row=7, column=2).value == "339:38"   # 9600+10778 minutes
    assert ws.cell(row=7, column=11).value == 41
    assert ws.cell(row=7, column=1).font.color.rgb.endswith("CC0000")
    # rate columns empty on every line
    assert ws.cell(row=5, column=3).value in ("", None)
    assert ws.cell(row=7, column=6).value in ("", None)
