"""
Step 11 — reinforcement guards (מתגברים): one-off external helpers on the
actual board, fully outside the attendance/payroll world.
"""

import io
from datetime import datetime, time, timedelta

import pytest
from openpyxl import load_workbook

import app.attendance.dependencies as attendance_deps
from app.attendance.dependencies import build_comparison_service
from app.exceptions import ConflictException
from app.models.user import User
from app.repositories.schedule_week_repository import ScheduleWeekRepository
from app.repositories.submission_repository import SubmissionRepository
from app.repositories.user_repository import UserRepository
from app.services.excel_export_service import ExcelExportService

from tests.test_actual_export_parity import (
    _actual_export,
    _make_rich_week,
    _planned_export,
)
from tests.test_actual_positions_api import _planned_week, _service


class _FlagOn:
    ACTUAL_SCHEDULE_ENABLED = True


@pytest.fixture
def flag_on(monkeypatch):
    monkeypatch.setattr(attendance_deps, "get_settings", lambda: _FlagOn())


@pytest.mark.asyncio
async def test_card_creation_and_assignment(db_session):
    week, _, _ = await _planned_week(db_session)
    service = _service(db_session)

    card = await service.add_reinforcement(
        week.id, first_name="חיצוני", last_name="מתגבר", note="מחברת אבטחה ב",
    )
    assert card.user.is_reinforcement is True
    assert card.user.phone_number.startswith("EXT-")  # unique placeholder

    board = await service.get_board(week.id)
    (row,) = board["rows"]
    await service.assign(row["position_id"], 0, card.user_id)

    board = await service.get_board(week.id)
    assert len(board["reinforcements"]) == 1
    assert any(a.user_id == card.user_id for a in board["assignments"])


@pytest.mark.asyncio
async def test_reinforcement_is_invisible_to_team_lists(db_session):
    week, _, _ = await _planned_week(db_session)
    await _service(db_session).add_reinforcement(
        week.id, first_name="חיצוני", last_name="מתגבר",
    )

    repo = UserRepository(db_session)
    assert all(not u.is_reinforcement for u in await repo.get_all_users())
    assert all(not u.is_reinforcement for u in await repo.get_active_users())


@pytest.mark.asyncio
async def test_reinforcement_never_reaches_attendance(db_session, flag_on):
    """Assigned but never punching — must NOT appear as a no-show anywhere."""
    week, _, _ = await _planned_week(db_session)
    service = _service(db_session)
    card = await service.add_reinforcement(
        week.id, first_name="חיצוני", last_name="מתגבר",
    )
    board = await service.get_board(week.id)
    (row,) = board["rows"]
    await service.assign(row["position_id"], 0, card.user_id)

    comparison = await build_comparison_service(db_session)
    now = datetime.combine(week.end_date + timedelta(days=1), time(12, 0))
    day_all = await comparison.get_day_all(week.start_date, now=now)
    assert all(r.user_id != card.user_id for r in day_all["rows"])

    # by_guard (the comparison/payroll feed) skips him; the grid cut keeps him.
    schedule = await _actual_export(db_session).get_week_schedule(week.id)
    assert all(g.user_id != card.user_id for g in schedule.by_guard)
    placements = [
        p.user_name
        for r in schedule.by_position for d in r.days for p in d.placements
    ]
    assert "חיצוני מתגבר" in placements


@pytest.mark.asyncio
async def test_reinforcement_shows_in_actual_excel(db_session):
    week = await _make_rich_week(db_session)
    service = _service(db_session)
    card = await service.add_reinforcement(
        week.id, first_name="חיצוני", last_name="מתגבר",
    )
    board = await service.get_board(week.id)
    arnona = next(r for r in board["rows"] if r["name"] == "ארנונה")
    await service.assign(arnona["position_id"], 5, card.user_id)

    excel = ExcelExportService(
        SubmissionRepository(db_session), UserRepository(db_session),
        ScheduleWeekRepository(db_session),
        _planned_export(db_session), _actual_export(db_session),
    )
    data = await excel.export_actual_schedule_grid(week.id)
    ws = load_workbook(io.BytesIO(data)).active
    texts = " ".join(
        str(c.value) for r in ws.iter_rows() for c in r if c.value is not None
    )
    assert "חיצוני מתגבר" in texts


@pytest.mark.asyncio
async def test_remove_card_deletes_user_and_assignments(db_session):
    week, _, _ = await _planned_week(db_session)
    service = _service(db_session)
    card = await service.add_reinforcement(
        week.id, first_name="חיצוני", last_name="מתגבר",
    )
    card_id, user_id = card.id, card.user_id
    board = await service.get_board(week.id)
    (row,) = board["rows"]
    await service.assign(row["position_id"], 0, user_id)

    await service.remove_reinforcement(card_id)

    board = await service.get_board(week.id)
    assert board["reinforcements"] == []
    assert board["assignments"] == []
    assert await UserRepository(db_session).get_by_id(user_id) is None


@pytest.mark.asyncio
async def test_duplicate_phone_rejected_kindly(db_session):
    week, _, _ = await _planned_week(db_session)
    existing = User(
        phone_number="0507777777", first_name="קיים", last_name="בצוות",
    )
    db_session.add(existing)
    await db_session.commit()

    with pytest.raises(ConflictException):
        await _service(db_session).add_reinforcement(
            week.id, first_name="חיצוני", last_name="מתגבר",
            phone_number="0507777777",
        )
