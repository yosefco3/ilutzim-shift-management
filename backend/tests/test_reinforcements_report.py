"""
Step 13 — the reinforcements report ("דוח מתגברים"): names, work dates and
hours over any period, with the supervisor field on the card.
"""

import io
from datetime import timedelta

import pytest
from openpyxl import load_workbook

from app.constants import WeekStatus
from app.exceptions import ValidationException
from app.models.schedule_week import ScheduleWeek
from app.models.user import User
from app.schedule_builder.models.activation_profile import ActivationProfile
from app.schedule_builder.models.position import Position
from app.schedule_builder.models.week_profile_assignment import WeekProfileAssignment
from app.utils.date_utils import today_il

from tests.test_actual_positions_api import _service


def _sheet_rows(data: bytes) -> list[list]:
    ws = load_workbook(io.BytesIO(data)).active
    return [[c.value for c in row] for row in ws.iter_rows()]


def _texts(rows: list[list]) -> str:
    return " ".join(str(v) for row in rows for v in row if v is not None)


async def _week_with_position(db_session, *, weeks_ago: int, profile=None):
    start = today_il() - timedelta(days=today_il().weekday() + 1 + 7 * weeks_ago)
    week = ScheduleWeek(
        start_date=start, end_date=start + timedelta(days=6),
        status=WeekStatus.LOCKED,
    )
    db_session.add(week)
    await db_session.flush()
    if profile is None:
        profile = ActivationProfile(name=f"שגרה {weeks_ago}", is_default=weeks_ago == 1)
        db_session.add(profile)
        await db_session.flush()
    db_session.add(WeekProfileAssignment(week_id=week.id, profile_id=profile.id))
    position = Position(
        profile_id=profile.id, name="שער ראשי",
        day_schedules={str(d): {"start": "07:00", "end": "15:00"} for d in range(7)},
    )
    db_session.add(position)
    await db_session.commit()
    return week, profile


@pytest.mark.asyncio
async def test_report_spans_weeks_with_dates_hours_and_supervisor(db_session):
    service = _service(db_session)
    week1, profile = await _week_with_position(db_session, weeks_ago=2)
    week2, _ = await _week_with_position(db_session, weeks_ago=1, profile=profile)

    card1 = await service.add_reinforcement(
        week1.id, first_name="חיצוני", last_name="אחד",
        phone_number="0501111222", supervisor_name="מפקח כהן",
    )
    card2 = await service.add_reinforcement(
        week2.id, first_name="חיצוני", last_name="שתיים",
    )

    board1 = await service.get_board(week1.id)
    row1 = next(r for r in board1["rows"] if r["name"] == "שער ראשי")
    # Day 1, a 4-hour segment.
    await service.assign(row1["position_id"], 1, card1.user_id, "07:00", "11:00")

    board2 = await service.get_board(week2.id)
    row2 = next(r for r in board2["rows"] if r["name"] == "שער ראשי")
    # Day 3, whole window (8h).
    await service.assign(row2["position_id"], 3, card2.user_id)

    # A regular team guard must NOT appear in the report.
    team_guard = User(phone_number="0509999888", first_name="צוות", last_name="אורגני")
    db_session.add(team_guard)
    await db_session.commit()
    await service.assign(row2["position_id"], 4, team_guard.id)

    start, end = week1.start_date, week2.end_date
    data = await service.export_reinforcements_report(start, end)
    rows = _sheet_rows(data)
    texts = _texts(rows)

    assert "חיצוני אחד" in texts and "חיצוני שתיים" in texts
    assert "צוות אורגני" not in texts
    assert "מפקח כהן" in texts and "0501111222" in texts
    date1 = (week1.start_date + timedelta(days=1)).strftime("%d/%m/%Y")
    date2 = (week2.start_date + timedelta(days=3)).strftime("%d/%m/%Y")
    assert date1 in texts and date2 in texts
    assert "07:00–11:00" in texts and "07:00–15:00" in texts
    assert "4:00" in texts and "8:00" in texts  # per-guard totals
    assert "12:00" in texts  # grand total 4h + 8h
    # Card without a supervisor / with a placeholder phone renders as "—".
    assert "EXT-" not in texts


@pytest.mark.asyncio
async def test_daily_cut_filters_to_one_date(db_session):
    service = _service(db_session)
    week, _ = await _week_with_position(db_session, weeks_ago=1)
    card = await service.add_reinforcement(
        week.id, first_name="חיצוני", last_name="יומי",
    )
    board = await service.get_board(week.id)
    (row,) = board["rows"]
    await service.assign(row["position_id"], 1, card.user_id)
    await service.assign(row["position_id"], 2, card.user_id)

    day = week.start_date + timedelta(days=1)
    data = await service.export_reinforcements_report(day, day)
    texts = _texts(_sheet_rows(data))
    assert day.strftime("%d/%m/%Y") in texts
    assert (week.start_date + timedelta(days=2)).strftime("%d/%m/%Y") not in texts


@pytest.mark.asyncio
async def test_night_shift_crossing_midnight_counts_fully(db_session):
    service = _service(db_session)
    week, _ = await _week_with_position(db_session, weeks_ago=1)
    night = await service.add_position(
        week.id, name="סייר לילה",
        day_schedules={"2": {"start": "23:00", "end": "07:00"}},
    )
    card = await service.add_reinforcement(
        week.id, first_name="חיצוני", last_name="לילה",
    )
    await service.assign(night.id, 2, card.user_id)

    day = week.start_date + timedelta(days=2)
    data = await service.export_reinforcements_report(day, day)
    texts = _texts(_sheet_rows(data))
    assert "23:00–07:00" in texts
    assert "8:00" in texts  # a full 8 hours, not negative/zero


@pytest.mark.asyncio
async def test_invalid_range_and_empty_period(db_session):
    service = _service(db_session)
    week, _ = await _week_with_position(db_session, weeks_ago=1)

    with pytest.raises(ValidationException):
        await service.export_reinforcements_report(
            week.end_date, week.start_date
        )

    data = await service.export_reinforcements_report(
        week.start_date, week.end_date
    )
    assert "אין שיבוצי מתגברים" in _texts(_sheet_rows(data))


@pytest.mark.asyncio
async def test_supervisor_returned_on_board(db_session):
    service = _service(db_session)
    week, _ = await _week_with_position(db_session, weeks_ago=1)
    await service.add_reinforcement(
        week.id, first_name="חיצוני", last_name="מפוקח",
        supervisor_name="מפקח לוי",
    )
    board = await service.get_board(week.id)
    (card,) = board["reinforcements"]
    assert card.supervisor_name == "מפקח לוי"
