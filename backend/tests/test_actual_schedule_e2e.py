"""
Step 10 — the full actual-schedule story, end to end.

One realistic week: rollover births the copy → day-of replacement → ad-hoc
event position → a position cancelled → punches recorded → with
ACTUAL_SCHEDULE_ENABLED the comparison, the payroll norm and the actual Excel
all reflect the edited reality — while the planning layer stays frozen.
Then: a retroactive fix on an ended week, and save-as-profile.
"""

import io
from datetime import datetime, time, timedelta

import pytest
from openpyxl import load_workbook

import app.attendance.dependencies as attendance_deps
from app.attendance.constants import ShiftPairStatus
from app.attendance.dependencies import build_comparison_service
from app.constants import WeekStatus
from app.models.schedule_week import ScheduleWeek
from app.models.user import User
from app.repositories.schedule_week_repository import ScheduleWeekRepository
from app.repositories.submission_repository import SubmissionRepository
from app.repositories.user_repository import UserRepository
from app.schedule_builder.models.activation_profile import ActivationProfile
from app.schedule_builder.models.position import Position
from app.schedule_builder.models.schedule_assignment import ScheduleAssignment
from app.schedule_builder.models.week_profile_assignment import WeekProfileAssignment
from app.schedule_builder.repositories.actual_schedule_repository import (
    ActualScheduleRepository,
)
from app.schedule_builder.repositories.position_repository import PositionRepository
from app.services.excel_export_service import ExcelExportService
from app.utils.date_utils import today_il

from tests.test_actual_export_parity import _actual_export, _planned_export
from tests.test_actual_positions_api import _service as _actual_service
from tests.test_attendance_comparison import _mk_shift


class _FlagOn:
    ACTUAL_SCHEDULE_ENABLED = True


@pytest.fixture
def flag_on(monkeypatch):
    monkeypatch.setattr(attendance_deps, "get_settings", lambda: _FlagOn())


def _xlsx_texts(data: bytes) -> str:
    ws = load_workbook(io.BytesIO(data)).active
    return " ".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    )


async def _build_planned_week(db_session):
    """A locked (started) week: two positions, two guards, full plan."""
    start = today_il() - timedelta(days=today_il().weekday() + 8)
    week = ScheduleWeek(
        start_date=start, end_date=start + timedelta(days=6),
        status=WeekStatus.LOCKED,
    )
    db_session.add(week)
    await db_session.flush()

    profile = ActivationProfile(name="שגרה", is_default=True)
    db_session.add(profile)
    await db_session.flush()
    db_session.add(WeekProfileAssignment(week_id=week.id, profile_id=profile.id))

    gate = Position(
        profile_id=profile.id, name="שער ראשי",
        day_schedules={str(d): {"start": "07:00", "end": "15:00"} for d in range(6)},
        display_order=1,
    )
    patrol = Position(
        profile_id=profile.id, name="סייר",
        day_schedules={"4": {"start": "15:00", "end": "23:00"}},
        display_order=2,
    )
    db_session.add_all([gate, patrol])
    await db_session.flush()

    planned_guard = User(phone_number="0501111111", first_name="אלף", last_name="מתוכנן")
    substitute = User(phone_number="0502222222", first_name="בית", last_name="מחליף")
    db_session.add_all([planned_guard, substitute])
    await db_session.flush()

    db_session.add_all([
        ScheduleAssignment(week_id=week.id, position_id=gate.id, day_index=2,
                           user_id=planned_guard.id),
        ScheduleAssignment(week_id=week.id, position_id=patrol.id, day_index=4,
                           user_id=planned_guard.id),
    ])
    await db_session.commit()
    return week, gate, patrol, planned_guard, substitute


@pytest.mark.asyncio
async def test_the_full_story(db_session, flag_on):
    week, gate, patrol, planned_guard, substitute = await _build_planned_week(db_session)
    service = _actual_service(db_session)
    repo = ActualScheduleRepository(db_session)

    # ── 1. The copy is born (here: lazily — rollover path is covered in 02). ──
    actual = await service.ensure_for_week(week.id)
    positions = {p.name: p for p in await repo.list_positions(actual.id)}

    # ── 2. Day 2 (ג'): guard A cancels day-of — replaced by guard B. ──
    assignments = await repo.list_assignments(actual.id)
    day2 = next(a for a in assignments if a.day_index == 2)
    await service.unassign(day2.id)
    await service.assign(positions["שער ראשי"].id, 2, substitute.id)

    # ── 3. Day 3 (ד'): unforeseen event — an ad-hoc position, 4 hours. ──
    adhoc = await service.add_position(
        week.id, name="אבטחת אירוע במתנס",
        day_schedules={"3": {"start": "16:00", "end": "20:00"}},
    )
    await service.assign(adhoc.id, 3, substitute.id)

    # ── 4. Day 4 (ה'): the patrol position is cancelled outright. ──
    await service.remove_position(positions["סייר"].id)

    # ── 5. Reality punches in (guard B, both days, exact). ──
    d2 = week.start_date + timedelta(days=2)
    d3 = week.start_date + timedelta(days=3)
    await _mk_shift(db_session, substitute,
                    in_at=datetime.combine(d2, time(7, 0)),
                    out_at=datetime.combine(d2, time(15, 0)),
                    status=ShiftPairStatus.COMPLETE, work_date=d2)
    await _mk_shift(db_session, substitute,
                    in_at=datetime.combine(d3, time(16, 0)),
                    out_at=datetime.combine(d3, time(20, 0)),
                    status=ShiftPairStatus.COMPLETE, work_date=d3)

    now = datetime.combine(week.end_date + timedelta(days=1), time(12, 0))
    comparison = await build_comparison_service(db_session)

    # ── 6a. The comparison reads the edited reality. ──
    day2_all = await comparison.get_day_all(d2, now=now)
    b_row = next(r for r in day2_all["rows"] if r.user_id == substitute.id)
    assert b_row.planned and b_row.summary.severity == "ok"
    assert all(r.user_id != planned_guard.id for r in day2_all["rows"])

    day3_all = await comparison.get_day_all(d3, now=now)
    b_day3 = next(r for r in day3_all["rows"] if r.user_id == substitute.id)
    assert b_day3.summary.planned_minutes == 4 * 60  # the ad-hoc position
    assert b_day3.summary.severity == "ok"

    day4_all = await comparison.get_day_all(
        week.start_date + timedelta(days=4), now=now
    )
    assert all(  # the cancelled patrol produces no phantom no-show
        r.user_id != planned_guard.id for r in day4_all["rows"]
    )

    # ── 6b. The payroll norm follows the actual board. ──
    from app.attendance.services.payroll_readmodel import PayrollReadModel

    payroll = PayrollReadModel(comparison, UserRepository(db_session), comparison.config)
    # the relative test week may straddle a month boundary — fetch every
    # month that d2/d3 fall in, so the assertion holds on any run date
    norm_by_day = {}
    for year, mon in {(d.year, d.month) for d in (d2, d3)}:
        month = await payroll.get_month(substitute.id, year, mon, now=now)
        norm_by_day.update({r.day: r.norm_minutes for r in month.rows if r.norm_minutes})
    assert norm_by_day[d2] == 8 * 60
    assert norm_by_day[d3] == 4 * 60

    # ── 6c. The actual Excel shows the reality; the plan stays frozen. ──
    excel = ExcelExportService(
        SubmissionRepository(db_session), UserRepository(db_session),
        ScheduleWeekRepository(db_session),
        _planned_export(db_session), _actual_export(db_session),
    )
    actual_texts = _xlsx_texts(await excel.export_actual_schedule_grid(week.id))
    assert "אבטחת אירוע במתנס" in actual_texts
    assert "בית מחליף" in actual_texts
    assert "סייר" not in actual_texts  # cancelled

    planned_texts = _xlsx_texts(await excel.export_schedule_grid(week.id))
    assert "אבטחת אירוע במתנס" not in planned_texts
    assert "סייר" in planned_texts  # the plan still remembers it
    assert "אלף מתוכנן" in planned_texts

    # And the planning tables themselves are untouched.
    assert await PositionRepository(db_session).get_by_id(patrol.id) is not None

    # ── 7. Retroactive fix on the (long-ended) week still works. ──
    extra_guard = User(phone_number="0503333333", first_name="גימל", last_name="מאוחר")
    db_session.add(extra_guard)
    await db_session.commit()
    await service.assign(positions["שער ראשי"].id, 0, extra_guard.id)

    comparison = await build_comparison_service(db_session)  # fresh cache
    day0_all = await comparison.get_day_all(week.start_date, now=now)
    late_row = next(r for r in day0_all["rows"] if r.user_id == extra_guard.id)
    assert late_row.planned  # the retro edit reached the comparison

    # ── 8. Save-as-profile promotes the edited board (incl. ad-hoc row). ──
    profile = await service.save_as_profile(week.id, "שגרה + מתנס")
    cloned = await PositionRepository(db_session).get_by_profile(profile.id)
    assert {p.name for p in cloned} == {"שער ראשי", "אבטחת אירוע במתנס"}
