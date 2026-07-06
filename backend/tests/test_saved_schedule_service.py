"""
Tests for SavedScheduleService (part B — saved schedule snapshot).

The headline guarantee: a saved snapshot is self-contained and survives deletion
of the profile it was built from (profile delete cascades positions →
assignments, but the snapshot keeps position/guard names inline).
"""

import uuid
from datetime import date

import pytest

from app.constants import WeekStatus
from app.exceptions import WeekNotFoundException
from app.models.schedule_week import ScheduleWeek
from app.models.user import User
from app.repositories.schedule_week_repository import ScheduleWeekRepository
from app.schedule_builder.models.activation_profile import ActivationProfile
from app.schedule_builder.models.position import Position
from app.schedule_builder.repositories.assignment_repository import (
    AssignmentRepository,
)
from app.schedule_builder.repositories.position_repository import PositionRepository
from app.schedule_builder.repositories.profile_repository import ProfileRepository
from app.schedule_builder.repositories.saved_schedule_repository import (
    SavedScheduleRepository,
)
from app.schedule_builder.repositories.week_profile_repository import (
    WeekProfileRepository,
)
from app.schedule_builder.services.assignment_service import AssignmentService
from app.schedule_builder.services.board_service import BoardService
from app.schedule_builder.services.saved_schedule_service import (
    SavedScheduleService,
)
from app.schedule_builder.services.week_profile_service import WeekProfileService


# The board-edit gate freezes assignments once a week has started; these fixtures
# use a 2026-07-05 week, so pin "today" before it stays editable regardless of the
# real clock (would otherwise trip after 2026-07-05).
@pytest.fixture(autouse=True)
def _pin_board_edit_clock():
    from unittest.mock import patch
    from datetime import date as _date
    with patch(
        "app.schedule_builder.services.assignment_service.today_il",
        return_value=_date(2026, 7, 1),
    ):
        yield


# ── fixtures / helpers ────────────────────────────────────────────────

def _board_service(session) -> BoardService:
    return BoardService(
        ScheduleWeekRepository(session),
        WeekProfileService(
            WeekProfileRepository(session),
            ProfileRepository(session),
            ScheduleWeekRepository(session),
        ),
        PositionRepository(session),
    )


def _service(session) -> SavedScheduleService:
    return SavedScheduleService(
        SavedScheduleRepository(session),
        _board_service(session),
        AssignmentRepository(session),
        ScheduleWeekRepository(session),
    )


def _assignment_service(session) -> AssignmentService:
    return AssignmentService(
        AssignmentRepository(session),
        ScheduleWeekRepository(session),
        PositionRepository(session),
    )


async def _make_week(db_session, start=date(2026, 7, 5)):
    week = ScheduleWeek(
        start_date=start, end_date=date(2026, 7, 11), status=WeekStatus.OPEN
    )
    db_session.add(week)
    await db_session.flush()
    return week


async def _make_profile(db_session, name="שגרה", is_default=True):
    profile = ActivationProfile(name=name, is_default=is_default)
    db_session.add(profile)
    await db_session.flush()
    return profile


async def _add_position(db_session, profile, name="ארנונה", active_days=("0",)):
    pos = Position(
        profile_id=profile.id,
        name=name,
        day_schedules={d: {"start": "07:00", "end": "15:00"} for d in active_days},
    )
    db_session.add(pos)
    await db_session.flush()
    return pos


async def _make_user(db_session, phone="0501111111", first="ישראל", last="ישראלי"):
    user = User(
        phone_number=phone, first_name=first, last_name=last,
        roles=[], is_active=True,
    )
    db_session.add(user)
    await db_session.flush()
    return user


def _cell(snapshot, day_index=0, row=0):
    return snapshot["rows"][row]["cells"][day_index]


# ── tests ─────────────────────────────────────────────────────────────

class TestSavedScheduleService:
    async def test_unknown_week_raises(self, db_session):
        with pytest.raises(WeekNotFoundException):
            await _service(db_session).save(uuid.uuid4())

    async def test_save_builds_snapshot(self, db_session):
        profile = await _make_profile(db_session)
        pos = await _add_position(db_session, profile, name="ארנונה")
        week = await _make_week(db_session)
        user = await _make_user(db_session, first="ישראל", last="ישראלי")
        await _assignment_service(db_session).assign(week.id, pos.id, 0, user.id)

        saved = await _service(db_session).save(week.id)

        assert saved.week_id == week.id
        assert saved.profile_name == "שגרה"
        snap = saved.snapshot
        assert snap["rows"][0]["position_name"] == "ארנונה"
        placed = _cell(snap, day_index=0)["assignments"]
        assert placed == [
            {"guard_name": "ישראל ישראלי", "segment_start": None, "segment_end": None}
        ]

    async def test_save_captures_segment(self, db_session):
        profile = await _make_profile(db_session)
        pos = await _add_position(db_session, profile)
        week = await _make_week(db_session)
        user = await _make_user(db_session)
        await _assignment_service(db_session).assign(
            week.id, pos.id, 0, user.id, "19:00", "01:00"
        )

        snap = (await _service(db_session).save(week.id)).snapshot
        seg = _cell(snap, day_index=0)["assignments"][0]
        assert (seg["segment_start"], seg["segment_end"]) == ("19:00", "01:00")

    async def test_resave_overwrites_single_row(self, db_session):
        profile = await _make_profile(db_session)
        pos = await _add_position(db_session, profile)
        week = await _make_week(db_session)
        u1 = await _make_user(db_session, "0501111111", "אבי", "כהן")
        svc = _service(db_session)
        asvc = _assignment_service(db_session)

        a = await asvc.assign(week.id, pos.id, 0, u1.id)
        await svc.save(week.id)
        # Replace the guard, save again.
        await asvc.unassign(a.id)
        u2 = await _make_user(db_session, "0502222222", "דנה", "לוי")
        await asvc.assign(week.id, pos.id, 0, u2.id)
        await svc.save(week.id)

        rows = await SavedScheduleRepository(db_session).list_all()
        assert len(rows) == 1  # still one snapshot for the week
        names = [
            p["guard_name"]
            for p in _cell(rows[0].snapshot, day_index=0)["assignments"]
        ]
        assert names == ["דנה לוי"]

    async def test_snapshot_survives_profile_deletion(self, db_session):
        """The core guarantee: delete the profile (cascading positions →
        assignments) — the saved snapshot and its inline names remain."""
        profile = await _make_profile(db_session)
        pos = await _add_position(db_session, profile, name="ארנונה")
        week = await _make_week(db_session)
        user = await _make_user(db_session, first="ישראל", last="ישראלי")
        await _assignment_service(db_session).assign(week.id, pos.id, 0, user.id)
        await _service(db_session).save(week.id)

        # Delete the profile → positions + assignments cascade away. A fresh
        # query (not the identity-map cache) confirms the assignments are gone —
        # which proves the position cascade fired too.
        await ProfileRepository(db_session).delete(profile.id)
        await db_session.flush()
        assert await AssignmentRepository(db_session).list_for_week(week.id) == []

        # The snapshot still exists, fully intact.
        saved = await _service(db_session).get(week.id)
        assert saved is not None
        snap = saved.snapshot
        assert snap["rows"][0]["position_name"] == "ארנונה"
        assert (
            _cell(snap, day_index=0)["assignments"][0]["guard_name"]
            == "ישראל ישראלי"
        )


class TestRenderSavedSchedule:
    """The Excel render is pure (snapshot dict → bytes); it must produce a valid
    workbook containing the inline names even after the profile is gone."""

    async def test_render_produces_valid_xlsx_with_names(self, db_session):
        import io

        import openpyxl

        from app.services.excel_export_service import ExcelExportService

        profile = await _make_profile(db_session)
        pos = await _add_position(db_session, profile, name="ארנונה")
        week = await _make_week(db_session)
        user = await _make_user(db_session, first="ישראל", last="ישראלי")
        await _assignment_service(db_session).assign(week.id, pos.id, 0, user.id)
        saved = await _service(db_session).save(week.id)

        # Delete the profile — render must still work from the snapshot alone.
        await ProfileRepository(db_session).delete(profile.id)
        await db_session.flush()

        # render_saved_schedule ignores the repos, so None args are fine here.
        data = ExcelExportService(None, None, None).render_saved_schedule(
            saved.snapshot
        )
        assert isinstance(data, bytes) and len(data) > 0

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        texts = {
            str(c.value)
            for col in ws.iter_cols()
            for c in col
            if c.value is not None
        }
        joined = " ".join(texts)
        assert "ארנונה" in joined
        assert "ישראל ישראלי" in joined

    def test_render_empty_snapshot_raises(self):
        from app.services.excel_export_service import ExcelExportService

        with pytest.raises(ValueError):
            ExcelExportService(None, None, None).render_saved_schedule({})

    def test_render_shows_hours_only_when_exceptional(self):
        """A regular full-window shift shows just the name; a deviating-hours day,
        partial coverage, or a split cell prints the hours in the cell."""
        import io

        import openpyxl

        from app.services.excel_export_service import ExcelExportService

        snapshot = {
            "week": {"start_date": "2025-01-05", "end_date": "2025-01-11"},
            "profile_name": "בדיקה",
            "days": [{"index": i, "date": ""} for i in range(7)],
            "rows": [{
                "position_name": "ארנונה",
                "band": "morning",
                "canonical_window": {"start": "07:00", "end": "15:00"},
                "cells": [
                    # day 0 — regular full window → name only
                    {"day_index": 0, "active": True,
                     "window": {"start": "07:00", "end": "15:00"},
                     "assignments": [
                         {"guard_name": "רגיל", "segment_start": None,
                          "segment_end": None}]},
                    # day 1 — deviating window (10:00 start) → hours shown
                    {"day_index": 1, "active": True,
                     "window": {"start": "10:00", "end": "15:00"},
                     "assignments": [
                         {"guard_name": "חריג", "segment_start": None,
                          "segment_end": None}]},
                    # day 2 — split between two guards → hours shown for both
                    {"day_index": 2, "active": True,
                     "window": {"start": "07:00", "end": "15:00"},
                     "assignments": [
                         {"guard_name": "בוקר", "segment_start": "07:00",
                          "segment_end": "11:00"},
                         {"guard_name": "צהריים", "segment_start": "11:00",
                          "segment_end": "15:00"}]},
                    # day 3 — active but no guard → amber gap
                    {"day_index": 3, "active": True,
                     "window": {"start": "07:00", "end": "15:00"},
                     "assignments": []},
                    # day 4 — inactive → grey
                    {"day_index": 4, "active": False,
                     "window": None, "assignments": []},
                ],
            }],
        }

        data = ExcelExportService(None, None, None).render_saved_schedule(snapshot)
        ws = openpyxl.load_workbook(io.BytesIO(data)).active

        assert ws.cell(row=4, column=1).value == "ארנונה\n07:00–15:00"
        assert ws.cell(row=4, column=2).value == "רגיל"  # day 0 — name only
        day1 = ws.cell(row=4, column=3).value            # deviating hours
        assert "חריג" in day1 and "10:00–15:00" in day1
        day2 = ws.cell(row=4, column=4).value            # split → both with hours
        assert "בוקר" in day2 and "07:00–11:00" in day2
        assert "צהריים" in day2 and "11:00–15:00" in day2
        gap = ws.cell(row=4, column=5)                    # active + empty → amber
        grey = ws.cell(row=4, column=6)                   # inactive → grey + ✕
        assert not gap.value
        assert gap.fill.start_color.rgb.endswith("FFC000")
        assert grey.value == "✕"
        assert grey.fill.start_color.rgb.endswith("E7E6E6")
