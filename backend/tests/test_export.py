"""
Tests for Excel export service and controller.
"""

import io
import json
import uuid
from datetime import date, timedelta
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from app.constants import ShiftType, SubmissionStatus, WeekStatus
from app.messages import Messages
from app.services.excel_export_service import ExcelExportService


# ── Helpers ──────────────────────────────────────────────────────────


def _make_user(user_id=None, full_name="Test Guard", phone="0501234567"):
    user = MagicMock()
    user.id = user_id or uuid.uuid4()
    user.full_name = full_name
    user.phone_number = phone
    user.is_active = True
    return user


def _make_week(week_id=None, week_start=None):
    # Mirror the real ScheduleWeek attributes (start_date/end_date) so the
    # mock catches attribute-name drift. week_start/week_end kept as aliases
    # for the legacy weekly-grid export which still reads them.
    start = week_start or date(2025, 1, 5)  # Sunday
    end = start + timedelta(days=6)
    week = MagicMock()
    week.id = week_id or uuid.uuid4()
    week.start_date = start
    week.end_date = end
    week.week_start = start
    week.week_end = end
    week.status = WeekStatus.OPEN
    return week


def _make_submission(user_id, week_id, status=SubmissionStatus.SUBMITTED):
    sub = MagicMock()
    sub.id = uuid.uuid4()
    sub.user_id = user_id
    sub.week_id = week_id
    sub.status = status
    return sub


def _create_service(
    week=None,
    users=None,
    submissions=None,
):
    """Build an ExcelExportService with mocked repos."""
    sub_repo = AsyncMock()
    user_repo = AsyncMock()
    week_repo = AsyncMock()

    week_repo.get_by_id.return_value = week
    user_repo.get_active_users.return_value = users or []
    sub_repo.get_by_week.return_value = submissions or []
    sub_repo.get_submissions_for_week.return_value = submissions or []
    sub_repo.get_by_user.return_value = submissions or []
    user_repo.get_by_id.return_value = users[0] if users else None

    return ExcelExportService(
        submission_repo=sub_repo,
        user_repo=user_repo,
        week_repo=week_repo,
    )


# ── Weekly schedule export tests ────────────────────────────────────


@pytest.mark.asyncio
async def test_export_weekly_schedule_basic():
    """Basic weekly schedule export with one user and submission."""
    user = _make_user()
    week = _make_week()
    sub = _make_submission(user.id, week.id)

    svc = _create_service(week=week, users=[user], submissions=[sub])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_weekly_schedule(week.id)

    assert isinstance(data, bytes)
    assert len(data) > 0

    # Verify it's a valid Excel (ZIP magic bytes)
    assert data[:2] == b"PK"


@pytest.mark.asyncio
async def test_export_weekly_no_submission_auto_absence():
    """Users without submissions get auto-absence markers."""
    user = _make_user()
    week = _make_week()

    svc = _create_service(week=week, users=[user], submissions=[])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_weekly_schedule(week.id)

    assert isinstance(data, bytes)
    assert data[:2] == b"PK"


@pytest.mark.asyncio
async def test_export_weekly_week_not_found():
    """Raises ValueError for unknown week ID."""
    svc = _create_service()

    with pytest.raises(ValueError, match="not found"):
        await svc.export_weekly_schedule(uuid.uuid4())


@pytest.mark.asyncio
async def test_export_weekly_no_openpyxl():
    """Raises RuntimeError when openpyxl is not installed."""
    svc = _create_service(week=_make_week())

    with patch("app.services.excel_export_service.HAS_OPENPYXL", False):
        with pytest.raises(RuntimeError, match="openpyxl"):
            await svc.export_weekly_schedule(svc._week_repo.get_by_id.return_value.id)


# ── Constraints report tests ────────────────────────────────────────


def _make_shift_window(shift_type, start, end):
    sw = MagicMock()
    sw.shift_type = shift_type
    sw.start_time = start
    sw.end_time = end
    return sw


def _make_daily_status(day, is_available, shift_windows=None):
    ds = MagicMock()
    ds.date = day
    ds.is_available = is_available
    ds.shift_windows = shift_windows or []
    return ds


def _make_constraint_submission(user_id, week_id, daily_statuses=None, notes=None):
    sub = _make_submission(user_id, week_id)
    sub.daily_statuses = daily_statuses or []
    sub.general_notes = notes
    return sub


@pytest.mark.asyncio
async def test_export_constraints_basic():
    """Constraints report with one submitting guard."""
    from datetime import time

    user = _make_user(full_name="בני לוי", phone="0502222222")
    week = _make_week(week_start=date(2025, 1, 5))
    ds = _make_daily_status(
        date(2025, 1, 5),
        True,
        [_make_shift_window(ShiftType.MORNING, time(6, 0), time(14, 0))],
    )
    sub = _make_constraint_submission(
        user.id, week.id, [ds], notes="זמין רק בבקרים"
    )

    svc = _create_service(week=week, users=[user], submissions=[sub])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_constraints_report(week.id)

    assert isinstance(data, bytes)
    assert data[:2] == b"PK"


@pytest.mark.asyncio
async def test_export_constraints_week_not_found():
    """Raises ValueError for unknown week ID."""
    svc = _create_service()

    with pytest.raises(ValueError, match="not found"):
        await svc.export_constraints_report(uuid.uuid4())


@pytest.mark.asyncio
async def test_export_constraints_no_openpyxl():
    """Raises RuntimeError when openpyxl is not installed."""
    svc = _create_service(week=_make_week())

    with patch("app.services.excel_export_service.HAS_OPENPYXL", False):
        with pytest.raises(RuntimeError, match="openpyxl"):
            await svc.export_constraints_report(
                svc._week_repo.get_by_id.return_value.id
            )


@pytest.mark.asyncio
async def test_constraints_excel_has_shift_times_and_notes():
    """Verify the constraints Excel shows shift windows and notes."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    from datetime import time

    user = _make_user(full_name="בני לוי", phone="0502222222")
    week = _make_week(week_start=date(2025, 1, 5))
    ds_sun = _make_daily_status(
        date(2025, 1, 5),
        True,
        [
            _make_shift_window(ShiftType.NIGHT, time(22, 0), time(6, 0)),
            _make_shift_window(ShiftType.MORNING, time(6, 0), time(14, 0)),
        ],
    )
    ds_mon = _make_daily_status(date(2025, 1, 6), False)
    sub = _make_constraint_submission(
        user.id, week.id, [ds_sun, ds_mon], notes="הערה כללית"
    )

    svc = _create_service(week=week, users=[user], submissions=[sub])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_constraints_report(week.id)

    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active

    # RTL sheet for Hebrew
    assert ws.sheet_view.rightToLeft is True

    # Header row (row 4): id, name, phone, period, then days, then notes
    assert ws.cell(row=4, column=1).value == "מזהה"
    assert ws.cell(row=4, column=2).value == Messages.EXCEL_HEADER_NAME
    assert ws.cell(row=4, column=3).value == Messages.EXCEL_HEADER_PHONE
    assert ws.cell(row=4, column=4).value == "משמרת"
    assert ws.cell(row=4, column=12).value == "הערות"

    # Each guard spans three rows (בוקר / צהריים / ערב), starting at row 5.
    # The guard's DB id is injected in col 1, merged across the three rows.
    assert ws.cell(row=5, column=1).value == str(user.id)
    assert ws.cell(row=5, column=2).value == "בני לוי"
    assert ws.cell(row=5, column=4).value == "בוקר"
    assert ws.cell(row=6, column=4).value == "ערב"
    assert ws.cell(row=7, column=4).value == "לילה"

    # Sunday (col 5): morning window in the בוקר row, night window in the לילה row.
    assert "06:00" in ws.cell(row=5, column=5).value
    assert "22:00" in ws.cell(row=7, column=5).value
    # Afternoon (ערב) row has no window that day → empty.
    assert not ws.cell(row=6, column=5).value
    # Monday (col 6): not available → merged "לא זמין" on the top row.
    assert ws.cell(row=5, column=6).value == "לא זמין"
    # Notes merged across the three rows, anchored on the top row.
    assert ws.cell(row=5, column=12).value == "הערה כללית"


@pytest.mark.asyncio
async def test_constraints_excel_has_thick_separator_between_guards():
    """Each guard's three-row block ends with a heavy bottom border so
    adjacent guards are easy to tell apart."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    from datetime import time

    week = _make_week(week_start=date(2025, 1, 5))
    user_a = _make_user(full_name="אבי כהן", phone="0501111111")
    user_b = _make_user(full_name="בני לוי", phone="0502222222")
    ds_a = _make_daily_status(
        date(2025, 1, 5),
        True,
        [_make_shift_window(ShiftType.MORNING, time(7, 0), time(15, 0))],
    )
    ds_b = _make_daily_status(date(2025, 1, 5), False)
    sub_a = _make_constraint_submission(user_a.id, week.id, [ds_a])
    sub_b = _make_constraint_submission(user_b.id, week.id, [ds_b])

    svc = _create_service(
        week=week, users=[user_a, user_b], submissions=[sub_a, sub_b]
    )

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_constraints_report(week.id)

    ws = openpyxl.load_workbook(io.BytesIO(data)).active

    # Guards are sorted by name: אבי (rows 5-7), בני (rows 8-10).
    # openpyxl derives a merged range's bottom edge from its top-left (anchor)
    # cell, so the merged columns carry the thick side on the block's top row;
    # the non-merged period cells carry it on the block's last row.
    def _bottom(row, col):
        return ws.cell(row=row, column=col).border.bottom.style

    # Non-merged period column (4): thick on each block's last (לילה) row.
    assert _bottom(7, 4) == "thick"
    assert _bottom(10, 4) == "thick"
    # Merged id column (1): thick carried on the anchor row of each block.
    assert _bottom(5, 1) == "thick"
    assert _bottom(8, 1) == "thick"
    # Interior period rows are not separators.
    assert _bottom(5, 4) == "thin"
    assert _bottom(6, 4) == "thin"


@pytest.mark.asyncio
async def test_constraints_export_round_trips_through_import_parser():
    """The exported file (with the new מזהה column) must parse cleanly back
    through the constraints-import parser, recovering the guard's DB id."""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        pytest.skip("openpyxl not installed")

    from datetime import time
    from app.services.constraints_import.parser import parse_constraints_xlsx

    user = _make_user(full_name="בני לוי", phone="0502222222")
    week = _make_week(week_start=date(2025, 1, 5))
    ds_sun = _make_daily_status(
        date(2025, 1, 5),
        True,
        [_make_shift_window(ShiftType.MORNING, time(6, 0), time(14, 0))],
    )
    sub = _make_constraint_submission(
        user.id, week.id, [ds_sun], notes="הערה"
    )
    svc = _create_service(week=week, users=[user], submissions=[sub])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_constraints_report(week.id)

    parsed = parse_constraints_xlsx(data)

    # The extra ID column did not break parsing (no fatal header error).
    assert not any("פורמט" in e or "כותרת" in e for e in parsed.errors)
    assert len(parsed.guards) == 1
    g = parsed.guards[0]
    # The guard's DB id was injected into the file AND recovered by the parser.
    assert g.guard_id == str(user.id)
    assert g.name == "בני לוי"
    # Sunday morning window survived the round-trip.
    morning = g.cells[0][ShiftType.MORNING]
    assert morning.start == time(6, 0)
    assert morning.end == time(14, 0)


# ── Built-schedule grid tests (positions × days) ────────────────────


def _pos_row(name, band, placements_by_day, active_days=None, canonical=None,
             gaps_by_day=None, is_event=False, event_required_count=None):
    """Build a PositionRow with 7 day-columns. ``placements_by_day`` maps
    day_index → list of (name, start, end). ``active_days`` defaults to the days
    that carry placements (plus none extra). ``canonical`` is an optional
    ``(start, end)`` pair for the position's regular hours. ``gaps_by_day`` maps
    day_index → list of (start, end) uncovered gaps."""
    from app.schedule_builder.services.schedule_export_service import (
        Placement, PositionDay, PositionRow,
    )
    active = active_days if active_days is not None else set(placements_by_day)
    gaps_by_day = gaps_by_day or {}
    days = []
    for d in range(7):
        placed = [
            Placement(user_id=uuid.uuid4(), user_name=n, start=s, end=e)
            for (n, s, e) in placements_by_day.get(d, [])
        ]
        days.append(PositionDay(
            day_index=d, date="", active=(d in active or bool(placed)),
            placements=placed, gaps=list(gaps_by_day.get(d, [])),
        ))
    canon = {"start": canonical[0], "end": canonical[1]} if canonical else None
    return PositionRow(
        position_id=uuid.uuid4(), name=name, band=band, days=days,
        canonical_window=canon, is_event=is_event,
        event_required_count=event_required_count,
    )


def _grid_service(week, position_rows):
    """ExcelExportService wired to a mock schedule read model."""
    from app.schedule_builder.services.schedule_export_service import WeekSchedule

    week_repo = AsyncMock()
    week_repo.get_by_id.return_value = week
    schedule_export = AsyncMock()
    schedule_export.get_week_schedule.return_value = WeekSchedule(
        week=week, days=[{"index": i, "date": ""} for i in range(7)],
        by_position=position_rows, by_guard=[],
    )
    return ExcelExportService(
        submission_repo=AsyncMock(), user_repo=AsyncMock(),
        week_repo=week_repo, schedule_export_service=schedule_export,
    )


@pytest.mark.asyncio
async def test_schedule_grid_regular_shift_hides_hours():
    """A full-window shift on its regular hours shows just the name — the hours
    live under the position, not in every cell."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    week = _make_week(week_start=date(2025, 1, 5))
    rows = [_pos_row("ארנונה", "morning",
                     {0: [("ישראל ישראלי", "07:00", "15:00")]},
                     canonical=("07:00", "15:00"))]
    svc = _grid_service(week, rows)

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_schedule_grid(week.id)

    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    assert ws.sheet_view.rightToLeft is True
    assert ws.cell(row=3, column=1).value == "עמדה"
    assert ws.cell(row=3, column=2).value == "ראשון"
    # Position block starts at row 4; the regular hours sit under the name.
    assert ws.cell(row=4, column=1).value == "ארנונה\n07:00–15:00"
    sunday = ws.cell(row=4, column=2).value
    assert sunday == "ישראל ישראלי"  # name only — no repeated hours


@pytest.mark.asyncio
async def test_schedule_grid_deviating_shift_shows_hours():
    """A shift whose hours differ from the position's regular window (e.g. a
    Tuesday 10:00 start, or partial coverage) prints the hours in the cell."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    week = _make_week(week_start=date(2025, 1, 5))
    rows = [_pos_row("ארנונה", "morning", {
        0: [("ישראל ישראלי", "07:00", "15:00")],   # regular
        2: [("דנה לוי", "10:00", "15:00")],          # Tuesday deviation
    }, canonical=("07:00", "15:00"))]
    svc = _grid_service(week, rows)

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_schedule_grid(week.id)

    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    sunday = ws.cell(row=4, column=2).value   # regular → name only
    tuesday = ws.cell(row=4, column=4).value  # day_index 2 → col 4
    assert sunday == "ישראל ישראלי"
    assert "דנה לוי" in tuesday and "10:00–15:00" in tuesday


@pytest.mark.asyncio
async def test_schedule_grid_tiled_cell_splits_block():
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    week = _make_week(week_start=date(2025, 1, 5))
    rows = [_pos_row("ארנונה", "morning", {
        0: [("אבי כהן", "07:00", "13:00"), ("דנה לוי", "13:00", "19:00")],
    })]
    svc = _grid_service(week, rows)

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_schedule_grid(week.id)

    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    # The position block spans two sub-rows (rows 4 and 5); one guard each.
    first = ws.cell(row=4, column=2).value
    second = ws.cell(row=5, column=2).value
    assert "אבי כהן" in first and "07:00–13:00" in first
    assert "דנה לוי" in second and "13:00–19:00" in second


@pytest.mark.asyncio
async def test_schedule_grid_inactive_day_is_grey_and_survives():
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    week = _make_week(week_start=date(2025, 1, 5))
    # Active only on day 0; day 3 inactive.
    rows = [_pos_row("ארנונה", "morning",
                     {0: [("ישראל ישראלי", "07:00", "15:00")]}, active_days={0})]
    svc = _grid_service(week, rows)

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_schedule_grid(week.id)

    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    wednesday = ws.cell(row=4, column=5)  # day_index 3 → col 5
    assert wednesday.value == "✕"  # blocked-day mark
    assert wednesday.fill.start_color.rgb.endswith("E7E6E6")  # grey


@pytest.mark.asyncio
async def test_schedule_grid_partial_coverage_splits_with_amber_gap():
    """A partially-covered cell splits into the guard's part (name + hours) and
    an amber gap part (the uncovered hours, no name), ordered chronologically."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    week = _make_week(week_start=date(2025, 1, 5))
    # Guard covers 08:00–15:00; the 07:00–08:00 start is an uncovered gap.
    rows = [_pos_row("ארנונה", "morning",
                     {0: [("ישראל ישראלי", "08:00", "15:00")]},
                     canonical=("07:00", "15:00"),
                     gaps_by_day={0: [("07:00", "08:00")]})]
    svc = _grid_service(week, rows)

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_schedule_grid(week.id)

    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    # Block spans two sub-rows: gap (07:00–08:00) first, then the guard.
    gap = ws.cell(row=4, column=2)
    guard = ws.cell(row=5, column=2)
    assert gap.value == "07:00–08:00"
    assert gap.fill.start_color.rgb.endswith("FFC000")   # amber gap
    assert "ישראל ישראלי" in guard.value and "08:00–15:00" in guard.value


@pytest.mark.asyncio
async def test_schedule_grid_empty_active_cell_is_amber():
    """An active cell with no guard is flagged amber (a staffing gap), while an
    inactive day stays grey."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    week = _make_week(week_start=date(2025, 1, 5))
    # Days 0 and 1 active; guard only on day 0. Day 1 = active but unstaffed.
    rows = [_pos_row("ארנונה", "morning",
                     {0: [("ישראל ישראלי", "07:00", "15:00")]},
                     active_days={0, 1}, canonical=("07:00", "15:00"))]
    svc = _grid_service(week, rows)

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_schedule_grid(week.id)

    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    monday = ws.cell(row=4, column=3)     # day_index 1 → col 3, active + empty
    tuesday = ws.cell(row=4, column=4)    # day_index 2 → col 4, inactive
    assert not monday.value
    assert monday.fill.start_color.rgb.endswith("FFC000")   # amber gap
    assert tuesday.value == "✕"                             # blocked-day mark
    assert tuesday.fill.start_color.rgb.endswith("E7E6E6")  # grey inactive


@pytest.mark.asyncio
async def test_schedule_grid_single_guard_merges_across_split_block():
    """When one day in a row is split (two sub-rows), the other days' single-guard
    cells are merged down the whole block — only genuinely-split cells show as two
    rows, matching the board (not a name on top with an empty cell below)."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter  # noqa: F401
    except ImportError:
        pytest.skip("openpyxl not installed")

    week = _make_week(week_start=date(2025, 1, 5))
    rows = [_pos_row("ארנונה", "morning", {
        0: [("אבי כהן", "07:00", "13:00"), ("דנה לוי", "13:00", "19:00")],  # split → span 2
        2: [("שוהם לוי", "07:00", "15:00")],                                # lone full-window guard
    }, canonical=("07:00", "15:00"))]
    svc = _grid_service(week, rows)

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_schedule_grid(week.id)

    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    merged = {str(r) for r in ws.merged_cells.ranges}
    # The split Sunday stays two independent sub-rows…
    assert "אבי כהן" in ws.cell(row=4, column=2).value
    assert "דנה לוי" in ws.cell(row=5, column=2).value
    assert "B4:B5" not in merged
    # …while the lone-guard Tuesday (col 4) is merged across both rows, name only.
    assert ws.cell(row=4, column=4).value == "שוהם לוי"
    assert "D4:D5" in merged
    assert ws.cell(row=5, column=4).value is None  # part of the merge, no empty split


@pytest.mark.asyncio
async def test_schedule_grid_event_row_is_purple_and_never_amber():
    """An event (non-splitting) position is painted purple: its name block, its
    staffed cells, AND its empty active cells (an unstaffed event is a valid
    state, never an amber staffing hole). It also holds >2 guards in one cell."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    week = _make_week(week_start=date(2025, 1, 5))
    # Day 0: three guards on the same window (no tiling). Day 1: active + empty.
    rows = [_pos_row("רענון", "morning", {
        0: [("אבי כהן", "07:00", "15:00"),
            ("דנה לוי", "07:00", "15:00"),
            ("שוהם לוי", "07:00", "15:00")],
    }, active_days={0, 1}, canonical=("07:00", "15:00"), is_event=True)]
    svc = _grid_service(week, rows)

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_schedule_grid(week.id)

    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    # Name block is the deeper event purple.
    assert ws.cell(row=4, column=1).fill.start_color.rgb.endswith("CCC0DA")
    # Three guards stack in the Sunday cell (col 2) — no cap, no split.
    names = {ws.cell(row=4 + p, column=2).value for p in range(3)}
    assert any("אבי כהן" in (n or "") for n in names)
    assert any("שוהם לוי" in (n or "") for n in names)
    # Staffed event cell is light-purple, not white/amber.
    assert ws.cell(row=4, column=2).fill.start_color.rgb.endswith("E4DFEC")
    # Monday (day_index 1 → col 3) active but unstaffed → purple, NOT amber.
    monday = ws.cell(row=4, column=3)
    assert not monday.value
    assert monday.fill.start_color.rgb.endswith("E4DFEC")


@pytest.mark.asyncio
async def test_schedule_grid_empty_event_row_is_omitted():
    """An event position with no guards all week is dropped from the sheet; a
    staffed event (even a single guard) stays."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    week = _make_week(week_start=date(2025, 1, 5))
    rows = [
        _pos_row("רענון ריק", "morning", {},
                 active_days={0, 1}, canonical=("07:00", "15:00"), is_event=True),
        _pos_row("רענון מאויש", "morning",
                 {0: [("אבי כהן", "07:00", "15:00")]},
                 canonical=("07:00", "15:00"), is_event=True),
    ]
    svc = _grid_service(week, rows)

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_schedule_grid(week.id)

    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    names = [ws.cell(row=r, column=1).value for r in range(4, ws.max_row + 1)]
    joined = "\n".join(n for n in names if n)
    assert "רענון ריק" not in joined      # empty event dropped
    assert "רענון מאויש" in joined         # staffed event kept


@pytest.mark.asyncio
async def test_schedule_grid_fixed_count_event_shows_missing_slots():
    """A fixed-count event (4 required, 3 assigned) tiles into 4 slots: three
    guard names + one amber 'חסר' hole, so the admin sees a guard is missing."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    week = _make_week(week_start=date(2025, 1, 5))
    rows = [_pos_row("מועצה", "morning", {
        0: [("אבי כהן", "07:00", "15:00"),
            ("דנה לוי", "07:00", "15:00"),
            ("שוהם לוי", "07:00", "15:00")],
    }, active_days={0}, canonical=("07:00", "15:00"),
        is_event=True, event_required_count=4)]
    svc = _grid_service(week, rows)

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_schedule_grid(week.id)

    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    # The block is 4 rows tall (rows 4..7) — one per required participant.
    col = 2  # Sunday
    values = [ws.cell(row=4 + p, column=col).value for p in range(4)]
    names = [v for v in values if v and v != "חסר"]
    assert len(names) == 3
    # The 4th (missing) slot is an amber "חסר" hole.
    assert values[3] == "חסר"
    assert ws.cell(row=7, column=col).fill.start_color.rgb.endswith("FFC000")
    # Filled slots stay event-purple.
    assert ws.cell(row=4, column=col).fill.start_color.rgb.endswith("E4DFEC")


@pytest.mark.asyncio
async def test_schedule_grid_fixed_count_event_empty_day_stays_purple():
    """A fixed-count event day with NO guards (the event isn't happening that
    day) is a single purple cell — not N amber 'חסר' holes — while another day
    of the same row is fully staffed."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    week = _make_week(week_start=date(2025, 1, 5))
    rows = [_pos_row("מועצה", "morning", {
        0: [("אבי כהן", "07:00", "15:00"),
            ("דנה לוי", "07:00", "15:00")],
        # Day 1 active but no guards → event not happening that day.
    }, active_days={0, 1}, canonical=("07:00", "15:00"),
        is_event=True, event_required_count=2)]
    svc = _grid_service(week, rows)

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_schedule_grid(week.id)

    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    # Monday (day_index 1 → col 3): one merged purple cell, no "חסר" markers.
    monday = [ws.cell(row=4 + p, column=3).value for p in range(2)]
    assert not any(v == "חסר" for v in monday)
    assert ws.cell(row=4, column=3).fill.start_color.rgb.endswith("E4DFEC")


@pytest.mark.asyncio
async def test_schedule_grid_week_not_found():
    svc = _grid_service(week=None, position_rows=[])
    with pytest.raises(ValueError, match="not found"):
        await svc.export_schedule_grid(uuid.uuid4())


def _export_client(fake_service):
    """A TestClient over the export router with admin auth + service overridden."""
    from fastapi import FastAPI
    from fastapi.testclient import TestClient

    from app.controllers.admin_export_controller import router as export_router
    from app.dependencies import get_excel_export_service, require_admin_role

    app = FastAPI()
    app.include_router(export_router)
    app.dependency_overrides[require_admin_role] = lambda: None
    app.dependency_overrides[get_excel_export_service] = lambda: fake_service
    return TestClient(app)


def _guard_sched(name, shifts, telegram_id=None):
    """Build a GuardSchedule. ``shifts`` = list of (day_index, position, start, end)."""
    from app.schedule_builder.services.schedule_export_service import (
        GuardSchedule, GuardShift,
    )
    return GuardSchedule(
        user_id=uuid.uuid4(), user_name=name, telegram_id=telegram_id,
        shifts=[
            GuardShift(
                day_index=d, date="", position_id=uuid.uuid4(),
                position_name=p, start=s, end=e,
            )
            for (d, p, s, e) in shifts
        ],
    )


def _guard_positions_service(week, guards):
    from app.schedule_builder.services.schedule_export_service import WeekSchedule

    week_repo = AsyncMock()
    week_repo.get_by_id.return_value = week
    schedule_export = AsyncMock()
    schedule_export.get_week_schedule.return_value = WeekSchedule(
        week=week, days=[{"index": i, "date": ""} for i in range(7)],
        by_position=[], by_guard=guards,
    )
    return ExcelExportService(
        submission_repo=AsyncMock(), user_repo=AsyncMock(),
        week_repo=week_repo, schedule_export_service=schedule_export,
    )


@pytest.mark.asyncio
async def test_guard_positions_shows_each_day():
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    week = _make_week(week_start=date(2025, 1, 5))
    guard = _guard_sched("אבי כהן", [
        (0, "ארנונה", "07:00", "15:00"),
        (2, "קומה 6", "15:00", "23:00"),
        (4, "סייר", "23:00", "07:00"),
    ], telegram_id="tg-avi")
    svc = _guard_positions_service(week, [guard])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_guard_positions(week.id)

    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    assert ws.sheet_view.rightToLeft is True
    assert ws.cell(row=3, column=1).value == "שם"
    assert ws.cell(row=4, column=1).value == "אבי כהן"
    # Sunday (col 2), Tuesday (col 4), Thursday (col 6).
    assert "ארנונה" in ws.cell(row=4, column=2).value
    assert "קומה 6" in ws.cell(row=4, column=4).value
    assert "סייר" in ws.cell(row=4, column=6).value


@pytest.mark.asyncio
async def test_guard_positions_unscheduled_guard_gets_note():
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    week = _make_week(week_start=date(2025, 1, 5))
    guard = _guard_sched("דנה לוי", [], telegram_id="tg-dana")  # no shifts
    svc = _guard_positions_service(week, [guard])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_guard_positions(week.id)

    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    assert ws.cell(row=4, column=1).value == "דנה לוי"
    assert ws.cell(row=4, column=2).value == "אין שיבוצים השבוע"


@pytest.mark.asyncio
async def test_guard_positions_unverified_first_and_flagged():
    """A guard with no linked Telegram floats to the top of the sheet and their
    name block carries the "no Telegram" flag; linked guards stay plain."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    week = _make_week(week_start=date(2025, 1, 5))
    # Name order would put אבי first, but דנה has no Telegram → she leads.
    linked = _guard_sched(
        "אבי כהן", [(0, "ארנונה", "07:00", "15:00")], telegram_id="tg-avi"
    )
    unlinked = _guard_sched("דנה לוי", [(0, "ארנונה", "07:00", "15:00")])
    svc = _guard_positions_service(week, [linked, unlinked])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_guard_positions(week.id)

    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    top = ws.cell(row=4, column=1).value
    assert top.startswith("דנה לוי")
    assert "🚫 אין טלגרם" in top
    # The linked guard follows, unflagged.
    nxt = ws.cell(row=5, column=1).value
    assert nxt == "אבי כהן"


@pytest.mark.asyncio
async def test_guard_positions_two_positions_same_day_split_block():
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    week = _make_week(week_start=date(2025, 1, 5))
    guard = _guard_sched("אבי כהן", [
        (0, "ארנונה", "07:00", "15:00"),
        (0, "קומה 6", "15:00", "19:00"),
    ])
    svc = _guard_positions_service(week, [guard])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_guard_positions(week.id)

    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    # Block spans two sub-rows (4 and 5); both positions visible in Sunday col.
    first = ws.cell(row=4, column=2).value
    second = ws.cell(row=5, column=2).value
    assert "ארנונה" in first
    assert "קומה 6" in second


@pytest.mark.asyncio
async def test_guard_positions_week_not_found():
    svc = _guard_positions_service(week=None, guards=[])
    with pytest.raises(ValueError, match="not found"):
        await svc.export_guard_positions(uuid.uuid4())


def test_guard_positions_endpoint_200_and_404():
    class FakeExport:
        async def export_guard_positions(self, week_id):
            if str(week_id) == "00000000-0000-0000-0000-000000000000":
                raise ValueError(f"Week {week_id} not found")
            return b"PK\x03\x04fake-xlsx"

    client = _export_client(FakeExport())
    ok = client.get(f"/admin/export/guard-positions/{uuid.uuid4()}")
    assert ok.status_code == 200
    assert "spreadsheetml" in ok.headers["content-type"]
    missing = client.get(
        "/admin/export/guard-positions/00000000-0000-0000-0000-000000000000"
    )
    assert missing.status_code == 404


def test_schedule_grid_endpoint_200_and_404():
    from fastapi.testclient import TestClient  # noqa: F401

    class FakeExport:
        async def export_schedule_grid(self, week_id):
            if str(week_id) == "00000000-0000-0000-0000-000000000000":
                raise ValueError(f"Week {week_id} not found")
            return b"PK\x03\x04fake-xlsx"

    client = _export_client(FakeExport())

    ok = client.get(f"/admin/export/schedule/{uuid.uuid4()}")
    assert ok.status_code == 200
    assert "spreadsheetml" in ok.headers["content-type"]

    missing = client.get(
        "/admin/export/schedule/00000000-0000-0000-0000-000000000000"
    )
    assert missing.status_code == 404


def test_schedule_png_endpoint_200_and_404():
    class FakeExport:
        async def export_schedule_grid_png(self, week_id):
            if str(week_id) == "00000000-0000-0000-0000-000000000000":
                raise ValueError(f"Week {week_id} not found")
            return b"\x89PNG\r\n\x1a\nfake"

    client = _export_client(FakeExport())

    ok = client.get(f"/admin/export/schedule-png/{uuid.uuid4()}")
    assert ok.status_code == 200
    assert ok.headers["content-type"] == "image/png"

    missing = client.get(
        "/admin/export/schedule-png/00000000-0000-0000-0000-000000000000"
    )
    assert missing.status_code == 404


# ── Deviation report tests ──────────────────────────────────────────


@pytest.mark.asyncio
async def test_export_deviation_report_basic():
    """Deviation report with one guard below threshold."""
    user = _make_user()
    week = _make_week()
    sub = _make_submission(user.id, week.id)

    svc = _create_service(week=week, users=[user], submissions=[sub])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_deviation_report(week.id)

    assert isinstance(data, bytes)
    assert data[:2] == b"PK"


@pytest.mark.asyncio
async def test_export_deviation_missing_submission():
    """Guards with no submission are shown in deviation report."""
    user = _make_user()
    week = _make_week()

    svc = _create_service(week=week, users=[user], submissions=[])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_deviation_report(week.id)

    assert isinstance(data, bytes)


@pytest.mark.asyncio
async def test_export_deviation_week_not_found():
    """Raises ValueError for unknown week ID."""
    svc = _create_service()

    with pytest.raises(ValueError, match="not found"):
        await svc.export_deviation_report(uuid.uuid4())


# ── Guard history report tests ──────────────────────────────────────


@pytest.mark.asyncio
async def test_export_guard_history_basic():
    """Guard history report across a date range."""
    user = _make_user()
    start = date(2025, 1, 5)
    end = date(2025, 1, 19)  # 2 weeks

    svc = _create_service(users=[user], submissions=[])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_guard_history(user.id, start, end)

    assert isinstance(data, bytes)
    assert data[:2] == b"PK"


@pytest.mark.asyncio
async def test_export_guard_history_user_not_found():
    """Raises ValueError for unknown user ID."""
    svc = _create_service(users=[])

    with pytest.raises(ValueError, match="not found"):
        await svc.export_guard_history(uuid.uuid4(), date(2025, 1, 5), date(2025, 1, 19))


# ── Integration: verify Excel content via openpyxl ───────────────────


@pytest.mark.asyncio
async def test_weekly_excel_has_correct_structure():
    """Verify the generated Excel has correct title and header rows."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    user = _make_user(full_name="יוסי כהן")
    week = _make_week(week_start=date(2025, 1, 5))
    sub = _make_submission(user.id, week.id)

    svc = _create_service(week=week, users=[user], submissions=[sub])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_weekly_schedule(week.id)

    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active

    # Title row
    assert ws.cell(row=1, column=1).value is not None

    # Header row (row 3)
    assert ws.cell(row=3, column=1).value == Messages.EXCEL_HEADER_NAME
    assert ws.cell(row=3, column=2).value == "ראשון"

    # Data row (row 4) - user name
    assert ws.cell(row=4, column=1).value == "יוסי כהן"


@pytest.mark.asyncio
async def test_deviation_excel_has_correct_data():
    """Verify deviation Excel shows guards below threshold."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    user = _make_user(full_name="דני לוי", phone="0509876543")
    week = _make_week()
    sub = _make_submission(user.id, week.id)

    svc = _create_service(week=week, users=[user], submissions=[sub])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_deviation_report(week.id)

    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active

    # Header row
    assert ws.cell(row=3, column=1).value == Messages.EXCEL_HEADER_NAME

    # Data row - should have user with submission status
    assert ws.cell(row=4, column=1).value == "דני לוי"
