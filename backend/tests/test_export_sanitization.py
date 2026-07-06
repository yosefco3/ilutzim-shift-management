"""
Tests for Excel formula-injection sanitization in the export service.

User-controlled text (guard names, general notes, imported position names) is
written into Excel cells. A value like ``=HYPERLINK("http://evil","x")`` would
become a live formula when the admin opens the file. ``_sanitize_cell`` prefixes
a leading apostrophe to neutralize the dangerous prefixes (= + - @ \\t \\r).
"""

import io
import uuid
from datetime import date, time

import pytest

from app.constants import ShiftType
from app.services.excel_export_service import _sanitize_cell

from tests.test_export import (
    _create_service,
    _make_constraint_submission,
    _make_daily_status,
    _make_shift_window,
    _make_user,
    _make_week,
)


# ── unit: _sanitize_cell ─────────────────────────────────────────────────────

def test_sanitize_prefixes_dangerous_leading_chars():
    assert _sanitize_cell("=1+2") == "'=1+2"
    assert _sanitize_cell("+A1") == "'+A1"
    assert _sanitize_cell("-x") == "'-x"
    assert _sanitize_cell("@x") == "'@x"
    assert _sanitize_cell("\tx") == "'\tx"
    assert _sanitize_cell("\rx") == "'\rx"


def test_sanitize_leaves_safe_values_untouched():
    assert _sanitize_cell("יוסי כהן") == "יוסי כהן"
    assert _sanitize_cell("") == ""
    assert _sanitize_cell(None) is None
    assert _sanitize_cell(5) == 5
    # Internal Hebrew literals must not gain a stray apostrophe.
    assert _sanitize_cell("לא זמין") == "לא זמין"


# ── end-to-end: constraints report ───────────────────────────────────────────

@pytest.mark.asyncio
async def test_constraints_report_sanitizes_name_and_notes():
    """A malicious guard name / notes come back as text, not formulas."""
    import openpyxl

    evil_name = '=HYPERLINK("http://evil","x")'
    user = _make_user(full_name=evil_name, phone="0502222222")
    week = _make_week(week_start=date(2025, 1, 5))
    ds = _make_daily_status(
        date(2025, 1, 5),
        True,
        [_make_shift_window(ShiftType.MORNING, time(6, 0), time(14, 0))],
    )
    sub = _make_constraint_submission(user.id, week.id, [ds], notes="=1+2")

    svc = _create_service(week=week, users=[user], submissions=[sub])
    data = await svc.export_constraints_report(week.id)

    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active

    name_cells = [
        c.value for row in ws.iter_rows() for c in row
        if isinstance(c.value, str) and c.value.endswith('HYPERLINK("http://evil","x")')
    ]
    assert name_cells, "expected the guard name to be present"
    assert all(v == "'" + evil_name for v in name_cells)

    notes_cells = [
        c.value for row in ws.iter_rows() for c in row
        if isinstance(c.value, str) and c.value.endswith("1+2")
    ]
    assert notes_cells, "expected the notes cell to be present"
    assert all(v == "'=1+2" for v in notes_cells)

    # Blanket scan: no string cell may start with = / @ / tab.
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str):
                assert not c.value.startswith(("=", "@", "\t", "\r"))
